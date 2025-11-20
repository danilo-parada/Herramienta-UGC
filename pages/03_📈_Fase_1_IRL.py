import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import plotly.graph_objects as go
from pathlib import Path
from html import escape
import re
from typing import Any
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from core import irl_level_flow, trl, db, utils
from core.components import render_irl_banner
from core.theme import load_theme
from core.db_trl import save_trl_result, get_trl_history
from core.data_table import render_table

# Utilidades locales mínimas
def _clean_text(text: str | None) -> str:
    """Normaliza texto de evidencia: convierte None a '', recorta espacios y saltos."""
    return (text or "").strip()

# Banner informativo (placeholder seguro si no hay HTML específico)
IRL_IMPORTANT_HTML = (
    "<div class='irl-info'>Responde cada nivel. Si marcas VERDADERO, agrega antecedentes." "</div>"
)

def _safe_load_theme() -> None:
    """Carga el tema de forma tolerante a fallos.

    Evita NameError/ImportError en contextos donde el import tarde en resolverse o
    el runner recargue módulos fuera de orden.
    """
    try:
        from core.theme import load_theme as _lt  # type: ignore
        _lt()
    except Exception:
        # errores no críticos en inyección de tema no deben romper la página
        pass


def generate_irl_excel_template() -> bytes:
    """Genera plantilla Excel con todas las preguntas IRL para evaluación offline."""
    wb = Workbook()
    
    # Hoja 1: Instructivo
    ws_instructivo = wb.active
    ws_instructivo.title = "📋 Instructivo"
    
    instructivo_data = [
        ["INSTRUCCIONES PARA COMPLETAR LA EVALUACIÓN IRL"],
        [""],
        ["1. Vaya a la hoja 'Evaluación IRL'"],
        ["   TODAS las respuestas vienen PRE-LLENADAS con FALSO"],
        [""],
        ["2. Revise cada pregunta y CAMBIE a VERDADERO solo las que SÍ se cumplan:"],
        ["   • Si la pregunta NO aplica → Deje FALSO"],
        ["   • Si la pregunta SÍ aplica → Cambie a VERDADERO"],
        [""],
        ["3. Columna 'Evidencia':"],
        ["   • Si cambió a VERDADERO → Complete la evidencia (OBLIGATORIO)"],
        ["   • Si dejó en FALSO → Deje la celda vacía"],
        [""],
        ["4. IMPORTANTE - Escriba exactamente:"],
        ["   VERDADERO  (todo junto, sin espacios)"],
        ["   FALSO      (todo junto, sin espacios)"],
        [""],
        ["5. NO MODIFIQUE las columnas:"],
        ["   • Dimensión"],
        ["   • Nivel"],
        ["   • # Pregunta"],
        ["   • Pregunta"],
        [""],
        ["6. VENTAJA: Solo cambia las que SÍ cumplen, no pierde tiempo"],
        [""],
        ["7. Guarde el archivo (Ctrl+S) y súbalo en la plataforma"],
    ]
    
    for row in instructivo_data:
        ws_instructivo.append(row)
    
    # Estilos para instructivo
    for row in ws_instructivo.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True, size=14, color="FFFFFF")
            cell.fill = PatternFill(start_color="1b5e20", end_color="1b5e20", fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    ws_instructivo.column_dimensions['A'].width = 80
    
    # Hoja 2: Evaluación IRL
    ws_eval = wb.create_sheet("Evaluación IRL")
    
    # Encabezados
    headers = ["Dimensión", "Nivel", "# Pregunta", "Pregunta", "Respuesta", "Evidencia"]
    ws_eval.append(headers)
    
    # Estilo encabezados
    header_fill = PatternFill(start_color="2e7d32", end_color="2e7d32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws_eval[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Agregar todas las preguntas (sin fila de ejemplo)
    row_num = 2  # Empezar directamente después de encabezados
    for dimension, dim_desc in IRL_DIMENSIONS:
        levels = LEVEL_DEFINITIONS.get(dimension, [])
        
        for level in levels:
            level_id = level["nivel"]
            preguntas = level.get("preguntas", [])
            
            for idx, pregunta in enumerate(preguntas, start=1):
                ws_eval.append([
                    f"{dimension} - {dim_desc}",
                    level_id,
                    idx,
                    pregunta,
                    "FALSO",  # Pre-llenado con FALSO por defecto
                    ""   # Evidencia vacía
                ])
                row_num += 1
    
    # Ajustar anchos de columna
    ws_eval.column_dimensions['A'].width = 25
    ws_eval.column_dimensions['B'].width = 8
    ws_eval.column_dimensions['C'].width = 10
    ws_eval.column_dimensions['D'].width = 60
    ws_eval.column_dimensions['E'].width = 15
    ws_eval.column_dimensions['F'].width = 50
    
    # Aplicar bordes y alineación
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws_eval.iter_rows(min_row=2, max_row=row_num-1):
        for idx, cell in enumerate(row, 1):
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Forzar formato de TEXTO en columna E (Respuesta) para evitar conversión a booleano
            if idx == 5:  # Columna E (Respuesta)
                cell.number_format = '@'  # Formato de texto
    
    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def load_irl_excel_responses(uploaded_file) -> dict:
    """Carga respuestas desde Excel con mapeo inteligente y validación robusta."""
    try:
        df = pd.read_excel(uploaded_file, sheet_name="Evaluación IRL", header=0)
        
        # 🔍 PASO 1: Detectar columnas automáticamente
        col_dimension = None
        col_nivel = None
        col_num_pregunta = None
        col_pregunta = None
        col_respuesta = None
        col_evidencia = None
        
        for col in df.columns:
            col_lower = str(col).lower()
            if "dimensión" in col_lower or "dimension" in col_lower:
                col_dimension = col
            elif "nivel" in col_lower:
                col_nivel = col
            elif "#" in col_lower and "pregunta" in col_lower:
                col_num_pregunta = col
            elif "pregunta" in col_lower and "#" not in col_lower:
                col_pregunta = col
            elif "respuesta" in col_lower:
                col_respuesta = col
            elif "evidencia" in col_lower:
                col_evidencia = col
        
        # Mostrar columnas detectadas
        st.info(f"🔍 **Mapeo de columnas detectado:**\n- Dimensión: `{col_dimension}`\n- Nivel: `{col_nivel}`\n- # Pregunta: `{col_num_pregunta}`\n- Respuesta: `{col_respuesta}`\n- Evidencia: `{col_evidencia}`")
        
        # Validar que se encontraron todas las columnas necesarias
        if not all([col_dimension, col_nivel, col_num_pregunta, col_respuesta, col_evidencia]):
            st.error("❌ No se pudieron detectar todas las columnas necesarias")
            st.error(f"Columnas disponibles: {list(df.columns)}")
            return {}
        
        # 🔍 PASO 2: Procesar filas con validación detallada
        respuestas = {}
        stats = {
            'total': 0,
            'validas': 0,
            'invalidas': 0,
            'ejemplo': 0,
            'vacias': 0,
            'errores': []
        }
        
        for idx_row, row in df.iterrows():
            try:
                stats['total'] += 1
                
                # Leer dimensión
                dimension_full = str(row[col_dimension]) if pd.notna(row[col_dimension]) else ""
                if not dimension_full or dimension_full == 'nan':
                    stats['vacias'] += 1
                    continue
                
                dimension = dimension_full.split(" - ")[0].strip()
                
                # Leer nivel y número de pregunta
                if not pd.notna(row[col_nivel]) or not pd.notna(row[col_num_pregunta]):
                    stats['vacias'] += 1
                    continue
                
                level_id = int(row[col_nivel])
                pregunta_num = int(row[col_num_pregunta])
                
                # Leer y limpiar respuesta - IMPORTANTE: Excel convierte VERDADERO/FALSO a booleanos
                respuesta_raw = row[col_respuesta]
                
                # Manejar valores booleanos de Excel
                if isinstance(respuesta_raw, bool):
                    respuesta = "VERDADERO" if respuesta_raw else "FALSO"
                elif pd.notna(respuesta_raw):
                    respuesta_str = str(respuesta_raw).strip().upper().replace("\n", "").replace("\r", "").replace(" ", "").replace("\t", "")
                    
                    # Convertir variantes comunes
                    if respuesta_str in ["TRUE", "T", "1"]:
                        respuesta = "VERDADERO"
                    elif respuesta_str in ["FALSE", "F", "0"]:
                        respuesta = "FALSO"
                    elif respuesta_str in ["VERDADERO", "V"]:
                        respuesta = "VERDADERO"
                    elif respuesta_str in ["FALSO"]:
                        respuesta = "FALSO"
                    else:
                        respuesta = respuesta_str
                else:
                    respuesta = ""
                
                # DEBUG: Si es vacío o "NAN", registrar
                if not respuesta or respuesta == "NAN":
                    stats['vacias'] += 1
                    continue
                
                # Validar respuesta final
                if respuesta not in ["VERDADERO", "FALSO"]:
                    stats['invalidas'] += 1
                    stats['errores'].append({
                        'fila': idx_row + 2,
                        'dimension': dimension,
                        'nivel': level_id,
                        'pregunta_num': pregunta_num,
                        'respuesta_raw': f"'{respuesta_raw}' (tipo: {type(respuesta_raw).__name__})",
                        'respuesta_limpia': f"'{respuesta}'",
                        'motivo': f'No se pudo normalizar a VERDADERO/FALSO'
                    })
                    continue
                
                # Leer evidencia
                evidencia = str(row[col_evidencia]) if pd.notna(row[col_evidencia]) else ""
                
                # Crear keys en el formato correcto
                resp_key = f"resp_{dimension}_{level_id}_{pregunta_num}"
                toggle_key = f"toggle_{dimension}_{level_id}_{pregunta_num}"
                evid_key = f"evid_{dimension}_{level_id}_{pregunta_num}"
                
                respuestas[resp_key] = respuesta
                respuestas[toggle_key] = (respuesta == "VERDADERO")
                respuestas[evid_key] = evidencia.strip()
                
                stats['validas'] += 1
                
            except Exception as e:
                stats['invalidas'] += 1
                stats['errores'].append({
                    'fila': idx_row + 2,
                    'motivo': f'Error: {str(e)}'
                })
        
        # 📊 PASO 3: Mostrar reporte detallado
        st.markdown("---")
        st.markdown("### 📊 Reporte de Carga")
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("📄 Filas Totales", stats['total'])
        with col2:
            st.metric("✅ Válidas", stats['validas'], delta=f"{stats['validas']/stats['total']*100:.0f}%" if stats['total'] > 0 else "0%")
        with col3:
            st.metric("⚠️ Inválidas", stats['invalidas'])
        with col4:
            st.metric("📝 Omitidas", stats['ejemplo'] + stats['vacias'])
        
        # Mostrar errores si existen
        if stats['errores']:
            with st.expander(f"⚠️ Ver detalles de {len(stats['errores'])} fila(s) con problemas (primeras 30)"):
                for i, error in enumerate(stats['errores'][:30], 1):
                    if 'respuesta_raw' in error:
                        st.text(f"{i}. Fila {error['fila']}: {error['dimension']}-N{error['nivel']}-P{error['pregunta_num']}")
                        st.text(f"   Valor original: {error['respuesta_raw']}")
                        st.text(f"   Limpiado: {error['respuesta_limpia']}")
                        # Mostrar bytes para detectar caracteres invisibles
                        if error['respuesta_limpia'] != "''":
                            respuesta_bytes = error['respuesta_limpia'].strip("'").encode('utf-8')
                            st.text(f"   Bytes: {respuesta_bytes}")
                        st.text(f"   ❌ {error['motivo']}")
                    else:
                        st.text(f"{i}. Fila {error.get('fila', '?')}: {error['motivo']}")
                    st.text("")
        
        # Mensaje final
        if not respuestas:
            st.error("❌ No se encontraron respuestas válidas en el archivo.")
            return {}
        else:
            st.success(f"✅ Carga exitosa: {stats['validas']} respuestas listas para procesar")
        
        return respuestas
    
    except Exception as e:
        st.error(f"❌ Error crítico al leer el archivo Excel: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
        return {}

# Definiciones mínimas para restaurar el módulo tras un refactor: si alguna
# lista de niveles falta, definimos un contenedor vacío para mantener la app
# operativa. Las listas completas se definen más abajo cuando corresponda.
IRL_DIMENSIONS = [
    ("CRL", "Cliente"),
    ("BRL", "Modelo de negocio"),
    ("TRL", "Tecnología"),
    ("IPRL", "Propiedad Intelectual"),
    ("TmRL", "Equipo"),
    ("FRL", "Finanzas"),
]

# Algunas colecciones podrían no estar definidas en esta página; declaramos vacías por defecto
CRL_LEVELS: list[dict] = globals().get("CRL_LEVELS", [])
TRL_LEVELS: list[dict] = globals().get("TRL_LEVELS", [])
FRL_LEVELS: list[dict] = globals().get("FRL_LEVELS", [])

# Descripciones legibles por dimensión
DIMENSION_DESCRIPTIONS = {dim: desc for dim, desc in IRL_DIMENSIONS}


IPRL_LEVELS = [
    {
        "nivel": 1,
        "descripcion": "Se cuenta con una hipótesis sobre posibles derechos de propiedad intelectual que se podrían obtener (como patentes, software, derechos de autor, diseños, secretos comerciales, etc).",
        "preguntas": [
            "¿Tiene una hipótesis sobre posibles derechos de propiedad intelectual que se podrían obtener (como patentes, software, derechos de autor, diseños, secretos comerciales, etc.)?",
            "¿Tiene descripción y documentación de los posibles derechos de propiedad intelectual?",
            "¿Tiene claridad sobre aspectos legales relevantes o pertinentes (propiedad, derechos de uso, etc.)?",
            "¿Tiene conocimiento de los elementos únicos del invento y el campo técnico, estado del arte, publicaciones, etc.?",
        ],
    },
    {
        "nivel": 2,
        "descripcion": "Identificación de las diferentes formas de posibles derechos de propiedad intelectual que podrían tener. La propiedad de los derechos es clara y no hay dudas de ser el dueño de los derechos de PI",
        "preguntas": [
            "¿Ha mapeado las diferentes formas de derechos de propiedad intelectual que existen o podrían surgir durante el desarrollo?",
            "¿Tiene ideas específicas sobre los derechos de propiedad intelectual, aunque no estén bien descritas ni definidas?",
            "¿Ha identificado acuerdos relacionados con la propiedad intelectual y aclarado la propiedad?",
            "¿Ha identificado a los inventores/creadores y tiene conocimiento de las políticas de PI aplicables y potenciales restricciones en los contratos?",
        ],
    },
    {
        "nivel": 3,
        "descripcion": "Descripción detallada de los posibles derechos de propiedad intelectual claves (por ejemplo, invención o código).",
        "preguntas": [
            "¿Ha considerado qué formas de derechos de propiedad intelectual son claves o más importantes y podrían/deberían protegerse?",
            "¿Tiene una descripción suficientemente detallada de los posibles derechos de propiedad intelectual para evaluar la posibilidad de protección?",
            "¿Ha realizado una evaluación de las posibilidades de protección a través de búsquedas de publicaciones, estado del arte, soluciones de última generación, etc.?",
            "¿Ha realizado búsquedas o análisis iniciales del estado de la técnica pertinente o derechos de propiedad intelectual en conflicto con profesionales?",
        ],
    },
    {
        "nivel": 4,
        "descripcion": "Confirmación sobre la viabilidad de la protección y mediante qué mecanismo. Decisión sobre el por qué de proteger determinados derechos de propiedad intelectual (relevancia para el negocio).",
        "preguntas": [
            "¿Ha confirmado la viabilidad de la protección de los derechos de propiedad intelectual claves a través de búsquedas/análisis por parte de un profesional?",
            "¿Ha analizado los derechos de propiedad intelectual claves y definido prioridades sobre qué proteger para crear valor para el negocio/proyecto?",
            "¿Ha presentado la primera solicitud/registro de derechos de propiedad intelectual en una forma menos elaborada (por ejemplo, patente provisional)?",
        ],
    },
    {
        "nivel": 5,
        "descripcion": "Borrador de estrategia de los derechos de propiedad intelectual para usar estos derechos con fines comerciales. Presentación de la primera solicitud de patente completa.",
        "preguntas": [
            "¿Tiene un borrador de estrategia de los derechos de propiedad intelectual definida, idealmente por un profesional, sobre cómo usar los derechos de PI para proteger y ser valiosos para el negocio?",
            "¿Ha presentado la primera solicitud/registro formal completo de derechos de propiedad intelectual claves en cooperación con un profesional?",
            "¿Tiene acuerdos básicos vigentes para determinar el control de los derechos de propiedad intelectual claves (por ejemplo, asignaciones, propiedad, etc.)?",
        ],
    },
    {
        "nivel": 6,
        "descripcion": "La estrategia de protección se encuentra implementada y apoya el negocio. Respuesta positiva en solicitudes presentadas. Evaluación inicial de la libertad para operar.",
        "preguntas": [
            "¿Ha elaborado una estrategia completa de protección de los derechos de propiedad intelectual que sustenta la estrategia de negocio?",
            "¿Ha identificado posibles derechos de propiedad intelectual complementarios/adicionales a proteger?",
            "¿Ha realizado una evaluación inicial de la libertad para operar (freedom to operate) para comprender el panorama de los derechos de PI en el campo?",
            "¿Ha recibido respuesta positiva a las solicitudes de derechos de PI por parte de las autoridades?",
            "Si no ha recibido respuesta positiva, ¿ha realizado un análisis junto con profesionales con buenas perspectivas?",
        ],
    },
    {
        "nivel": 7,
        "descripcion": "Todos los derechos de propiedad intelectual claves han sido solicitados en los paises o regiones relevantes de acuerdo con la estrategia de derechos de propiedad intelectual",
        "preguntas": [
            "¿Ha solicitado todos los derechos de propiedad intelectual claves en los países o regiones relevantes de acuerdo con la estrategia de PI?",
            "¿Ha realizado una evaluación más completa de la libertad para operar y tiene una comprensión clara de la dependencia/restricción de otros derechos de PI existentes?",
        ],
    },
    {
        "nivel": 8,
        "descripcion": "Estrategia de protección y gestión de la propiedad intelectual completamente implementada. Evaluación más completa de la libertad de operar",
        "preguntas": [
            "¿Tiene una estrategia de protección y gestión de la propiedad intelectual completamente implementada?",
            "¿Ha sido otorgado los derechos de propiedad intelectual clave en el primer país/región con alcance relevante para el negocio?",
            "¿Ha presentado solicitud(es)/registro(s) de derechos de PI complementarios o adicionales?",
        ],
    },
    {
        "nivel": 9,
        "descripcion": "Sólido sustento y protección de derechos de propiedad intelectual para el negocio. Patente concedida y vigente en países relevantes",
        "preguntas": [
            "¿La estrategia de derechos de propiedad intelectual respalda y crea valor para el negocio?",
            "¿Se han otorgado y se mantienen los derechos de propiedad intelectual claves y complementarios en varios países relevantes para los negocios?",
            "¿Tiene acuerdos vigentes para acceder a todos los derechos de propiedad intelectual externos necesarios?",
        ],
    },
]

TMRL_LEVELS = [
    {
        "nivel": 1,
        "descripcion": "Poca comprensión de la necesidad de un equipo (generalmente un individuo). Falta de competencias y/o recursos necesarios.",
        "preguntas": [
            "¿El equipo está conformado por más de una persona que posee las competencias necesarias en áreas claves como tecnología y negocios?",
            "¿Tiene algo de conocimiento sobre las competencias y otros recursos necesarios (socios, proveedores de servicios, etc.) para verificar y desarrollar la idea?",
        ],
    },
    {
        "nivel": 2,
        "descripcion": "Conocimiento y primera idea sobre las competencias necesarias o los recursos externos (por ejemplo, socios) requeridos",
        "preguntas": [
            "¿Tiene una primera idea de qué personas/competencias adicionales podrían ser necesarias para verificar/desarrollar la idea?",
            "¿Tiene una primera idea del objetivo general del proyecto?",
        ],
    },
    {
        "nivel": 3,
        "descripcion": "Algunas de las competencias o recursos necesarios están presentes. Existen otras competencias o recursos que se necesitan y deben definirse (junto a un plan de búsqueda).",
        "preguntas": [
            "¿Existen personas en el equipo con algunas, pero no todas, las competencias necesarias para comenzar a verificar la idea?",
            "¿Ha identificado necesidades y brechas en competencias, capacidades y diversidad de equipos?",
            "¿Tiene un plan inicial sobre cómo encontrar las competencias necesarias a corto plazo (<1 año)?",
        ],
    },
    {
        "nivel": 4,
        "descripcion": "Un champion está presente. Varias de las competencias necesarias están presentes. Se inicia un plan para reclutar o asegurar recursos claves adicionales.",
        "preguntas": [
            "¿Hay un champion (impulsor y comprometido) en el equipo?",
            "¿El equipo tiene varias, pero no todas, las competencias necesarias, generalmente en múltiples individuos?",
            "¿Ha iniciado un plan para encontrar competencias y capacidades adicionales necesarias, teniendo en cuenta la diversidad del equipo?",
            "¿El equipo ha iniciado discusiones sobre roles, compromiso, propiedad, etc., para avanzar en el proyecto?",
        ],
    },
    {
        "nivel": 5,
        "descripcion": "El equipo fundador inicial ya posee las principales competencias necesarias. El equipo acuerda la propiedad y los roles, y tiene objetivos alineados",
        "preguntas": [
            "¿Existe un equipo fundador inicial trabajando juntos y dedicando un tiempo significativo al proyecto?",
            "¿El equipo fundador tiene en conjunto las principales competencias y capacidades necesarias para comenzar a construir la startup?",
            "¿El equipo está alineado con roles claros, metas y visiones compartidas y un claro compromiso con el proyecto?",
            "¿El equipo ha acordado sus respectivas participaciones accionarias con un acuerdo firmado?",
            "¿Se han iniciado actividades para obtener competencias y capacidades adicionales, teniendo en cuenta la diversidad del equipo?",
            "¿Se han implementado sistemas/procesos/herramientas iniciales para compartir conocimientos e información dentro del equipo?",
        ],
    },
    {
        "nivel": 6,
        "descripcion": "Existe un equipo complementario, diverso y comprometido, con todas las competencias y recursos necesarios, tanto en el ámbito de los negocios como el tecnológico.",
        "preguntas": [
            "¿Existe un equipo fundador complementario y diverso, capaz de comenzar a construir un negocio?",
            "¿Se cuenta con todas las competencias clave y la capacidad necesaria para el corto plazo, con claridad sobre quién es el director ejecutivo?",
            "¿El equipo está comprometido, todos sienten responsabilidad y están preparados para asumir responsabilidades?",
            "¿Ha iniciado la contratación de asesores y/o miembros del directorio, teniendo en cuenta la diversidad del directorio?",
            "¿Existe conciencia de los riesgos que pueden afectar el desempeño del equipo (conflictos, burnout/salud mental, política, etc.)?",
        ],
    },
    {
        "nivel": 7,
        "descripcion": "El equipo y la cultura de la empresa están plenamente establecidos y desarrollados de forma proactiva. Hay un plan visualizado para formar el equipo que se necesita a largo plazo",
        "preguntas": [
            "¿El equipo funciona bien con roles claros?",
            "¿Los objetivos, la visión, el propósito y la cultura están claramente articuladas y documentadas para apoyar al equipo y el desarrollo organizacional?",
            "¿Está en marcha un plan para desarrollar la organización y hacer crecer el equipo a largo plazo (~2 años)?",
            "¿Se han implementado procesos/sistemas y un plan de aprendizaje continuo para el desarrollo del personal?",
            "¿El Directorio y los asesores están en funcionamiento y apoyan al desarrollo empresarial y organizacional?",
        ],
    },
    {
        "nivel": 8,
        "descripcion": "Se cuenta con un CEO y equipo ejecutivo. Uso profesional del Directorio y de asesores. Se han activado planes y reclutamiento para la construcción de equipo a largo plazo.",
        "preguntas": [
            "¿Existe un liderazgo claro y un equipo de gestión con experiencia profesional relevante?",
            "¿Se cuenta con un Directorio competente y diverso, y asesores relevantes utilizados profesionalmente?",
            "¿Se han implementado políticas y procesos para asegurar buenas prácticas de recursos humanos y diversidad del equipo?",
            "¿Se están realizando contrataciones necesarias de acuerdo con el plan a largo plazo para determinar las competencias, capacidad y diversidad relevantes?",
            "¿Todos los niveles de la organización están debidamente capacitados y motivados?",
        ],
    },
    {
        "nivel": 9,
        "descripcion": "El equipo y la organización son de alto rendimiento y están correctamente estructurados. Ambos se mantienen y se desarrollan correctamente a lo largo del tiempo",
        "preguntas": [
            "¿La organización tiene un alto rendimiento y buen funcionamiento (cooperación, entorno social, etc.)?",
            "¿Todos los niveles de la organización participan activamente en el aprendizaje y el desarrollo continuo?",
            "¿La cultura organizacional, la estructura y los procesos se mejoran y desarrollan continuamente?",
            "¿Los incentivos/recompensas están alineados para motivar a toda la organización para alcanzar las metas y desempeñarse bien?",
            "¿El equipo directivo se mantiene, se desarrolla y se desempeña en el tiempo?",
        ],
    },
]

BRL_LEVELS = [
    {
        "nivel": 1,
        "descripcion": "Hipótesis preliminar sobre el concepto de negocio con información limitada del mercado.",
        "preguntas": [
            "¿Tiene una hipótesis preliminar del concepto de negocio?",
            "¿Cuenta con alguna información sobre el mercado y su potencial o tamaño?",
            "¿Tiene algún conocimiento o percepción de la competencia y soluciones alternativas?",
        ],
    },
    {
        "nivel": 2,
        "descripcion": "Descripción inicial estructurada del concepto de negocio y reconocimiento general del mercado.",
        "preguntas": [
            "¿Ha propuesto una descripción estructurada del concepto de negocio y la propuesta de valor?",
            "¿Se ha familiarizado brevemente con el tamaño del mercado, los segmentos y el panorama competitivo?",
            "¿Ha enumerado algunos competidores o alternativas?",
        ],
    },
    {
        "nivel": 3,
        "descripcion": "Borrador de modelo de negocios que caracteriza el mercado potencial y el panorama competitivo.",
        "preguntas": [
            "¿Ha generado un borrador del modelo de negocios (Canvas)?",
            "¿Ha descrito factores relevantes en el modelo de negocio que afectan al medio ambiente y la sociedad?",
            "¿Ha definido el mercado objetivo y estimado su tamaño (TAM, SAM)?",
            "¿Ha identificado y descrito la competencia y el panorama competitivo?",
        ],
    },
    {
        "nivel": 4,
        "descripcion": "Modelo de negocios completo inicial con primeras proyecciones de viabilidad económica.",
        "preguntas": [
            "¿Ha determinado la viabilidad económica a partir de las primeras proyecciones de pérdidas y ganancias?",
            "¿Ha realizado una evaluación inicial de la sostenibilidad ambiental y social?",
        ],
    },
    {
        "nivel": 5,
        "descripcion": "Modelo de negocios ajustado tras feedback de mercado y primeras hipótesis de ingresos.",
        "preguntas": [
            "¿Ha recibido feedback sobre los ingresos del modelo comercial de clientes potenciales o expertos?",
            "¿Ha recibido feedback sobre los costos del modelo comercial de socios, proveedores o expertos externos?",
            "¿Ha identificado medidas para aumentar las contribuciones ambientales y sociales positivas y disminuir las negativas?",
            "¿Ha actualizado la proyección de ganancias y pérdidas en función del feedback del mercado?",
            "¿Ha actualizado la descripción del mercado objetivo y el análisis competitivo basado en comentarios del mercado?",
        ],
    },
    {
        "nivel": 6,
        "descripcion": "Modelo de negocios sostenible validado mediante escenarios comerciales realistas.",
        "preguntas": [
            "¿Tiene un modelo de negocio sostenible probado en escenarios comerciales realistas (ventas de prueba, pedidos anticipados, pilotos, etc.)?",
            "¿Tiene proyecciones financieras completas basadas en comentarios de casos comerciales realistas?",
        ],
    },
    {
        "nivel": 7,
        "descripcion": "Product/market fit inicial con disposición de pago demostrada y proyecciones validadas.",
        "preguntas": [
            "¿Las primeras ventas/ingresos en términos comerciales demuestran la disposición a pagar de un número significativo de clientes?",
            "¿Existen proyecciones financieras completas validadas por primeras ventas/ingresos y datos?",
            "¿Tiene acuerdos vigentes con proveedores clave, socios y socios de canal alineados con sus expectativas de sostenibilidad?",
        ],
    },
    {
        "nivel": 8,
        "descripcion": "Modelo de negocios sostenible que demuestra capacidad de escalar con métricas operativas.",
        "preguntas": [
            "¿Las ventas y otras métricas de las operaciones comerciales iniciales muestran que el modelo de negocio sostenible se mantiene y puede escalar?",
            "¿Están establecidos y operativos los canales de venta y la cadena de suministro alineados con sus expectativas de sostenibilidad?",
            "¿El modelo comercial se ajusta para mejorar los ingresos/costos y aprovechar la sostenibilidad?",
        ],
    },
    {
        "nivel": 9,
        "descripcion": "Modelo de negocios definitivo y sostenible con ingresos recurrentes y métricas consolidadas.",
        "preguntas": [
            "¿El modelo de negocio es sostenible y operativo, y el negocio cumple o supera las expectativas internas y externas en cuanto a beneficios, crecimiento, escalabilidad e impacto ambiental y social?",
            "¿Utiliza sistemas y métricas creíbles para rastrear el desempeño económico, ambiental y social?",
            "¿Los datos históricos sobre el desempeño económico, ambiental y social prueban un negocio viable, rentable y sostenible en el tiempo?",
        ],
    },
]

# Dimensión Cliente (CRL)
CRL_LEVELS = [
    {
        "nivel": 1,
        "descripcion": "Hipótesis especulativa sobre una posible necesidad en el mercado.",
        "preguntas": [
            "¿Tiene alguna hipótesis sobre un problema o necesidad que podría existir en el mercado?",
            "¿Ha identificado quiénes podrían ser sus posibles clientes, aunque sea de manera especulativa?",
        ],
    },
    {
        "nivel": 2,
        "descripcion": "Familiarización inicial con el mercado y necesidades más específicas detectadas.",
        "preguntas": [
            "¿Ha realizado alguna investigación secundaria o revisión de mercado para entender problemas del cliente?",
            "¿Tiene una descripción más clara y específica de las necesidades o problemas detectados?",
        ],
    },
    {
        "nivel": 3,
        "descripcion": "Primer feedback de mercado y validación preliminar de necesidades.",
        "preguntas": [
            "¿Ha iniciado contactos directos con posibles usuarios o expertos del mercado para obtener retroalimentación?",
            "¿Ha comenzado a desarrollar una hipótesis más clara sobre los segmentos de clientes y sus problemas?",
        ],
    },
    {
        "nivel": 4,
        "descripcion": "Confirmación del problema con varios usuarios y segmentación inicial.",
        "preguntas": [
            "¿Ha confirmado el problema o necesidad con varios clientes o usuarios reales?",
            "¿Ha definido una hipótesis de producto basada en el feedback recibido de los usuarios?",
            "¿Tiene segmentación inicial de clientes en función del problema identificado?",
        ],
    },
    {
        "nivel": 5,
        "descripcion": "Interés establecido por parte de usuarios y comprensión más profunda del mercado.",
        "preguntas": [
            "¿Cuenta con evidencia de interés concreto por parte de clientes o usuarios hacia su solución?",
            "¿Ha establecido relaciones con potenciales clientes o aliados que retroalimentan su propuesta de valor?",
        ],
    },
    {
        "nivel": 6,
        "descripcion": "Beneficios de la solución confirmados a través de pruebas o asociaciones iniciales.",
        "preguntas": [
            "¿Ha realizado pruebas del producto o solución con clientes que validen sus beneficios?",
            "¿Ha iniciado procesos de venta o pilotos con clientes reales o aliados estratégicos?",
        ],
    },
    {
        "nivel": 7,
        "descripcion": "Clientes involucrados en pruebas extendidas o primeras ventas/test comerciales.",
        "preguntas": [
            "¿Tiene acuerdos o primeras ventas del producto (aunque sea versión de prueba)?",
            "¿Los clientes han participado activamente en validaciones o pruebas extendidas del producto?",
        ],
    },
    {
        "nivel": 8,
        "descripcion": "Ventas iniciales y preparación para ventas estructuradas y escalables.",
        "preguntas": [
            "¿Ha vendido sus primeros productos y validado la disposición de pago de un porcentaje relevante de clientes?",
            "¿Cuenta con una organización comercial mínima (CRM, procesos de venta, canales definidos)?",
        ],
    },
    {
        "nivel": 9,
        "descripcion": "Adopción consolidada y ventas repetibles a múltiples clientes reales.",
        "preguntas": [
            "¿Está realizando ventas escalables y repetibles con múltiples clientes?",
            "¿Su empresa está enfocada en ejecutar un proceso de crecimiento comercial con foco en la demanda de clientes?",
        ],
    },
]

# Dimensión Tecnología (TRL)
TRL_LEVELS = [
    {
        "nivel": 1,
        "descripcion": "Principios básicos observados y reportados.",
        "preguntas": [
            "¿Existe una descripción del principio científico/técnico subyacente?",
            "¿Se identificó el problema técnico a resolver?",
        ],
    },
    {
        "nivel": 2,
        "descripcion": "Formulación del concepto y/o aplicación tecnológica.",
        "preguntas": [
            "¿Se formuló el concepto de solución tecnológica?",
            "¿Existen hipótesis sobre mecanismos de funcionamiento y límites?",
        ],
    },
    {
        "nivel": 3,
        "descripcion": "Pruebas analíticas y experimentación de concepto en laboratorio.",
        "preguntas": [
            "¿Hay evidencia experimental preliminar que respalde el concepto?",
            "¿Se definieron métricas técnicas de éxito (KPIs técnicos)?",
        ],
    },
    {
        "nivel": 4,
        "descripcion": "Validación de componentes en laboratorio.",
        "preguntas": [
            "¿Los componentes clave fueron validados de forma aislada?",
            "¿Se cuenta con protocolos de ensayo reproducibles?",
        ],
    },
    {
        "nivel": 5,
        "descripcion": "Validación de componentes integrados en ambiente relevante.",
        "preguntas": [
            "¿El subsistema integrado cumple con los KPIs técnicos en ambiente relevante?",
            "¿Se identificaron riesgos técnicos críticos y mitigaciones?",
        ],
    },
    {
        "nivel": 6,
        "descripcion": "Demostración de sistema/prototipo en ambiente relevante.",
        "preguntas": [
            "¿Existe un prototipo funcional probado con condiciones de uso relevantes?",
            "¿Se validó el desempeño frente a condiciones operacionales variables?",
        ],
    },
    {
        "nivel": 7,
        "descripcion": "Demostración de sistema/prototipo en ambiente operacional real.",
        "preguntas": [
            "¿Se ejecutaron pilotos en campo con usuarios reales?",
            "¿El desempeño técnico cumplió con los KPIs en operación real?",
        ],
    },
    {
        "nivel": 8,
        "descripcion": "Sistema completo y calificado.",
        "preguntas": [
            "¿El sistema está completo, integrado y calificado por terceros (QA/QAQ)?",
            "¿Se cuenta con documentación de ingeniería y manuales de operación?",
        ],
    },
    {
        "nivel": 9,
        "descripcion": "Sistema probado en ambiente operacional real y listo para despliegue.",
        "preguntas": [
            "¿La tecnología está en operación estable y mantenible con clientes reales?",
            "¿Existen métricas de confiabilidad/disponibilidad en producción?",
        ],
    },
]

# Dimensión Finanzas (FRL)
FRL_LEVELS = [
    {
        "nivel": 1,
        "descripcion": "Hipótesis preliminares de modelo financiero (ingresos/costos) sin validación.",
        "preguntas": [
            "¿Existe una hipótesis de fuentes de ingreso?",
            "¿Se identificaron categorías de costos principales?",
        ],
    },
    {
        "nivel": 2,
        "descripcion": "Primer borrador de proyección financiera simple.",
        "preguntas": [
            "¿Se elaboró un primer P&G mensual o anual simple?",
            "¿Se estimó CAPEX/OPEX inicial a alto nivel?",
        ],
    },
    {
        "nivel": 3,
        "descripcion": "Modelo financiero básico con supuestos explícitos.",
        "preguntas": [
            "¿Se documentaron supuestos clave (precios, churn, costos variables)?",
            "¿Se construyó un flujo de caja simple con 12-24 meses?",
        ],
    },
    {
        "nivel": 4,
        "descripcion": "Validación inicial de supuestos de ingresos y costos.",
        "preguntas": [
            "¿Se contrastaron supuestos con datos de mercado/proveedores?",
            "¿Se identificó el punto de equilibrio preliminar?",
        ],
    },
    {
        "nivel": 5,
        "descripcion": "Modelo financiero iterado con escenarios y sensibilidad.",
        "preguntas": [
            "¿Se modelaron escenarios (base, optimista, conservador)?",
            "¿Se analizó sensibilidad de variables críticas (precio, conversión, CAC)?",
        ],
    },
    {
        "nivel": 6,
        "descripcion": "Evidencia de tracción financiera temprana (ingresos/pedidos/pilotos pagos).",
        "preguntas": [
            "¿Existen ingresos o contratos que respalden supuestos?",
            "¿Se actualizó el modelo con datos reales y desviaciones?",
        ],
    },
    {
        "nivel": 7,
        "descripcion": "Estructura financiera para escalar (unit economics positivos/near-breakeven).",
        "preguntas": [
            "¿Los unit economics son positivos o cercanos a equilibrio?",
            "¿Hay políticas y controles financieros operando (cobranza, compras)?",
        ],
    },
    {
        "nivel": 8,
        "descripcion": "Gestión financiera madura con reporting periódico.",
        "preguntas": [
            "¿Se reporta periódicamente P&G, balance y flujo de caja?",
            "¿Se cuenta con auditoría o revisión externa cuando aplica?",
        ],
    },
    {
        "nivel": 9,
        "descripcion": "Resultados financieros sostenidos y escalables.",
        "preguntas": [
            "¿La empresa muestra crecimiento rentable y sostenido?",
            "¿Las métricas financieras están estabilizadas y soportan escalamiento?",
        ],
    },
]

STEP_TABS = [dimension for dimension, _ in IRL_DIMENSIONS]
LEVEL_DEFINITIONS = {
    "CRL": CRL_LEVELS,
    "BRL": BRL_LEVELS,
    "TRL": TRL_LEVELS,
    "IPRL": IPRL_LEVELS,
    "TmRL": TMRL_LEVELS,
    "FRL": FRL_LEVELS,
}

STEP_CONFIG = {
    "min_evidence_chars": 40,
    "soft_char_limit": 400,
    "max_char_limit": 600,
    "evidence_obligatoria_strict": False,
    "secuencia_flexible": True,
}

_STATE_KEY = "irl_stepper_state"
_ERROR_KEY = "irl_stepper_errors"
_BANNER_KEY = "irl_stepper_banner"
_CLOSE_EXPANDER_KEY = "irl_close_expander"
_READY_KEY = "irl_level_ready"
_EDIT_MODE_KEY = "irl_level_edit_mode"
_AUTO_SAVE_KEY = "irl_auto_save"
_QUESTION_PROGRESS_KEY = "irl_question_progress"
# Claves adicionales usadas para restauración/estilos
_RESTORE_ON_EDIT_KEY = "irl_restore_on_edit"
_PENDING_RESTORE_QUEUE_KEY = "irl_pending_restore_queue"

# Mapa de estados → clase CSS (solo para estilos visuales)
_STATUS_CLASS_MAP = {
    "Pendiente": "pending",
    "Respondido (en cálculo)": "attention",
    "Revisión requerida": "review",
    "Error": "error",
    "Completa": "complete",
    "Completo": "complete",
}

def _is_evidence_valid(texto: str | None) -> bool:
    """Valida evidencia: por defecto basta con que no esté vacía.

    Si STEP_CONFIG["evidence_obligatoria_strict"] es True, exige tamaño mínimo.
    """
    txt = _clean_text(texto)
    if not txt:
        return False
    if STEP_CONFIG.get("evidence_obligatoria_strict"):
        return len(txt) >= int(STEP_CONFIG.get("min_evidence_chars", 1))
    return True

def _ensure_question_progress(dimension: str, level_id: int, total_questions: int) -> dict:
    """Asegura y devuelve el progreso de preguntas para un nivel."""
    if _QUESTION_PROGRESS_KEY not in st.session_state:
        st.session_state[_QUESTION_PROGRESS_KEY] = {dim: {} for dim in STEP_TABS}
    if dimension not in st.session_state[_QUESTION_PROGRESS_KEY]:
        st.session_state[_QUESTION_PROGRESS_KEY][dimension] = {}
    progress = st.session_state[_QUESTION_PROGRESS_KEY][dimension].get(level_id)
    if not isinstance(progress, dict):
        progress = {}
    saved_map = progress.get("saved")
    if not isinstance(saved_map, dict):
        saved_map = {}
    valid_keys = {str(idx) for idx in range(1, max(total_questions, 0) + 1)}
    for key in list(saved_map.keys()):
        if key not in valid_keys:
            saved_map.pop(key, None)
    for idx in range(1, max(total_questions, 0) + 1):
        saved_map.setdefault(str(idx), False)
    progress["saved"] = saved_map
    active = progress.get("active", 0)
    if total_questions <= 0:
        active = 0
    else:
        active = max(0, min(int(active), total_questions - 1))
    progress["active"] = active
    st.session_state[_QUESTION_PROGRESS_KEY][dimension][level_id] = progress
    return progress

def _question_is_complete(resp: str | None, evidencia: str | None) -> bool:
    """Una pregunta está completa si:
    - resp == FALSO, o
    - resp == VERDADERO y la evidencia es válida.
    """
    if resp == "FALSO":
        return True
    if resp == "VERDADERO":
        return _is_evidence_valid(evidencia)
    return False

def _level_state(dimension: str, level_id: int) -> dict:
    """Obtiene/crea el estado de un nivel en session_state."""
    if _STATE_KEY not in st.session_state:
        st.session_state[_STATE_KEY] = {}
    if dimension not in st.session_state[_STATE_KEY]:
        st.session_state[_STATE_KEY][dimension] = {}
    if level_id not in st.session_state[_STATE_KEY][dimension]:
        st.session_state[_STATE_KEY][dimension][level_id] = {
            "respuesta": "FALSO",
            "respuestas_preguntas": {},
            "evidencia": "",
            "evidencias_preguntas": {},
            "estado": "Pendiente",
            "estado_auto": "Pendiente",
            "en_calculo": False,
            "marcado_revision": False,
        }
    return st.session_state[_STATE_KEY][dimension][level_id]

def _update_ready_flag(dimension: str, level_id: int) -> None:
    """Recalcula el flag ready para un nivel en base a los widgets actuales."""
    niveles = LEVEL_DEFINITIONS.get(dimension, [])
    level_data = next((lvl for lvl in niveles if lvl.get("nivel") == level_id), None)
    if _READY_KEY not in st.session_state:
        st.session_state[_READY_KEY] = {dim: {} for dim in STEP_TABS}
    if level_data is None:
        st.session_state[_READY_KEY].setdefault(dimension, {})[level_id] = False
        return
    preguntas = level_data.get("preguntas") or []
    if preguntas:
        listo = True
        for idx in range(1, len(preguntas) + 1):
            resp_key = f"resp_{dimension}_{level_id}_{idx}"
            evid_key = f"evid_{dimension}_{level_id}_{idx}"
            valor = st.session_state.get(resp_key)
            if valor not in {"VERDADERO", "FALSO"}:
                listo = False
                break
            if valor == "VERDADERO" and not _is_evidence_valid(st.session_state.get(evid_key)):
                listo = False
                break
    else:
        answer_key = f"resp_{dimension}_{level_id}"
        evidencia_key = f"evid_{dimension}_{level_id}"
        valor = st.session_state.get(answer_key)
        listo = valor in {"VERDADERO", "FALSO"}
        if listo and valor == "VERDADERO":
            listo = _is_evidence_valid(st.session_state.get(evidencia_key))
    st.session_state[_READY_KEY].setdefault(dimension, {})[level_id] = bool(listo)

def _rerun_app() -> None:
    try:
        st.rerun()
    except Exception:
        pass
def _render_level_question_flow(
    dimension: str,
    level_id: int,
    preguntas: list[str],
    descripcion: str,
    *,
    locked: bool,
) -> tuple[dict[str, str | None], dict[str, str], str, bool]:
    """Renderiza las preguntas del nivel sin botones extra.

    Devuelve las respuestas actuales, evidencias y si el nivel está listo para
    guardar. No crea botones; el botón único de guardado vive en el nivel.
    """

    irl_level_flow.inject_css()

    total_questions = len(preguntas)
    respuestas: dict[str, str | None] = {}
    evidencias: dict[str, str] = {}

    # Título del nivel
    st.markdown(f"### Nivel {level_id}")
    if descripcion:
        st.caption(descripcion)

    for idx, pregunta in enumerate(preguntas, start=1):
        resp_key = f"resp_{dimension}_{level_id}_{idx}"
        toggle_key = f"toggle_{dimension}_{level_id}_{idx}"
        evid_key = f"evid_{dimension}_{level_id}_{idx}"

        # Defaults seguros (solo si faltan)
        if resp_key not in st.session_state:
            st.session_state[resp_key] = "FALSO"
        if toggle_key not in st.session_state:
            st.session_state[toggle_key] = st.session_state[resp_key] == "VERDADERO"
        if evid_key not in st.session_state:
            st.session_state[evid_key] = ""

        st.write(f"**Pregunta {idx}/{total_questions}**: {pregunta}")

        # Widget del toggle
        selected = st.toggle(
            "VERDADERO/FALSO",
            key=toggle_key,
            disabled=locked,
        )
        # Sincronizamos la respuesta textual (no es widget)
        st.session_state[resp_key] = "VERDADERO" if selected else "FALSO"

        # Campo de evidencia (solo si VERDADERO). Importante: NO escribimos
        # sobre evid_key cuando el widget ya existe en esta misma ejecución.
        if selected:
            st.text_area(
                "Antecedentes de verificación",
                key=evid_key,
                disabled=locked,
                placeholder="Describe brevemente los antecedentes...",
                height=110,
            )
        else:
            # Si no se muestra el widget en esta ejecución es seguro limpiar
            # la evidencia para mantener consistencia
            if evid_key not in st.session_state or st.session_state.get(evid_key):
                st.session_state[evid_key] = ""

        respuestas[str(idx)] = st.session_state.get(resp_key)
        evidencias[str(idx)] = st.session_state.get(evid_key, "")

    evidencias_texto = "\n".join(
        t.strip() for t in evidencias.values() if isinstance(t, str) and t.strip()
    )

    ready_to_save = all(
        not (respuestas[str(i)] == "VERDADERO" and not (evidencias[str(i)] or "").strip())
        for i in range(1, total_questions + 1)
    )

    return respuestas, evidencias, evidencias_texto, ready_to_save
    if _QUESTION_PROGRESS_KEY not in st.session_state:
        st.session_state[_QUESTION_PROGRESS_KEY] = {dimension: {} for dimension in STEP_TABS}
    if dimension not in st.session_state[_QUESTION_PROGRESS_KEY]:
        st.session_state[_QUESTION_PROGRESS_KEY][dimension] = {}
    progress = st.session_state[_QUESTION_PROGRESS_KEY][dimension].get(level_id)
    if not isinstance(progress, dict):
        progress = {}
    saved_map = progress.get("saved")
    if not isinstance(saved_map, dict):
        saved_map = {}
    valid_keys = {str(idx) for idx in range(1, total_questions + 1)}
    for key in list(saved_map.keys()):
        if key not in valid_keys:
            saved_map.pop(key, None)
    for idx in range(1, total_questions + 1):
        saved_map.setdefault(str(idx), False)
    progress["saved"] = saved_map
    active = progress.get("active", 0)
    if active < 0 or active >= total_questions:
        active = 0
    progress["active"] = active
    st.session_state[_QUESTION_PROGRESS_KEY][dimension][level_id] = progress
    return progress


def _mark_question_pending(dimension: str, level_id: int, idx: int, total_questions: int) -> None:
    progress = _ensure_question_progress(dimension, level_id, total_questions)
    clave = str(idx)
    progress["saved"][clave] = False
    st.session_state[_QUESTION_PROGRESS_KEY][dimension][level_id] = progress
    error_key = f"question_error_{dimension}_{level_id}"
    if error_key in st.session_state:
        st.session_state[error_key] = None


def _mark_question_saved(dimension: str, level_id: int, idx: int, total_questions: int) -> None:
    progress = _ensure_question_progress(dimension, level_id, total_questions)
    clave = str(idx)
    progress["saved"][clave] = True
    st.session_state[_QUESTION_PROGRESS_KEY][dimension][level_id] = progress
    error_key = f"question_error_{dimension}_{level_id}"
    if error_key in st.session_state:
        st.session_state[error_key] = None


def _set_active_question(dimension: str, level_id: int, idx: int, total_questions: int) -> None:
    progress = _ensure_question_progress(dimension, level_id, total_questions)
    if total_questions <= 0:
        progress["active"] = 0
    else:
        progress["active"] = max(0, min(idx, total_questions - 1))
    st.session_state[_QUESTION_PROGRESS_KEY][dimension][level_id] = progress


def _init_irl_state() -> None:
    """Inicializa todas las estructuras necesarias en session_state."""
    # contenedores raíz
    if _STATE_KEY not in st.session_state:
        st.session_state[_STATE_KEY] = {}
    if _ERROR_KEY not in st.session_state:
        st.session_state[_ERROR_KEY] = {dim: {} for dim in STEP_TABS}
    if _BANNER_KEY not in st.session_state:
        st.session_state[_BANNER_KEY] = {}
    if _EDIT_MODE_KEY not in st.session_state:
        st.session_state[_EDIT_MODE_KEY] = {dim: {} for dim in STEP_TABS}
    if _READY_KEY not in st.session_state:
        st.session_state[_READY_KEY] = {dim: {} for dim in STEP_TABS}
    if _RESTORE_ON_EDIT_KEY not in st.session_state:
        st.session_state[_RESTORE_ON_EDIT_KEY] = {dim: {} for dim in STEP_TABS}
    if _QUESTION_PROGRESS_KEY not in st.session_state:
        st.session_state[_QUESTION_PROGRESS_KEY] = {dim: {} for dim in STEP_TABS}
    if _PENDING_RESTORE_QUEUE_KEY not in st.session_state:
        st.session_state[_PENDING_RESTORE_QUEUE_KEY] = []
    if "irl_scores" not in st.session_state:
        st.session_state["irl_scores"] = {dim: 0 for dim in STEP_TABS}

    # niveles por dimensión
    for dimension in STEP_TABS:
        st.session_state[_STATE_KEY].setdefault(dimension, {})
        st.session_state[_ERROR_KEY].setdefault(dimension, {})
        st.session_state[_EDIT_MODE_KEY].setdefault(dimension, {})
        st.session_state[_READY_KEY].setdefault(dimension, {})
        st.session_state[_RESTORE_ON_EDIT_KEY].setdefault(dimension, {})
        st.session_state[_QUESTION_PROGRESS_KEY].setdefault(dimension, {})

        niveles = LEVEL_DEFINITIONS.get(dimension, [])
        for level in niveles:
            nivel_id = level.get("nivel")
            # estado base del nivel
            state = _level_state(dimension, nivel_id)
            # normalizar respuestas por pregunta
            preguntas = level.get("preguntas") or []
            existentes = state.get("respuestas_preguntas")
            if not isinstance(existentes, dict):
                existentes = {}
            normalizado: dict[str, str | None] = {}
            for idx, _ in enumerate(preguntas, start=1):
                clave = str(idx)
                valor = existentes.get(clave)
                normalizado[clave] = valor if valor in {"VERDADERO", "FALSO"} else "FALSO"
            state["respuestas_preguntas"] = normalizado
            if not isinstance(state.get("evidencias_preguntas"), dict):
                state["evidencias_preguntas"] = {}

            # editar por defecto: True si no está calculado
            st.session_state[_EDIT_MODE_KEY][dimension].setdefault(nivel_id, not state.get("en_calculo", False))
            # errores inicialmente None
            st.session_state[_ERROR_KEY][dimension].setdefault(nivel_id, None)
            # flags de restauración
            st.session_state[_RESTORE_ON_EDIT_KEY][dimension].setdefault(nivel_id, False)
            # progreso preguntas
            _ensure_question_progress(dimension, nivel_id, len(preguntas))
            # ready flag
            _update_ready_flag(dimension, nivel_id)


def _set_level_state(
    dimension: str,
    level_id: int,
    *,
    respuesta: str | None = None,
    respuestas_preguntas: dict[str, str | None] | None = None,
    evidencia: str | None = None,
    evidencias_preguntas: dict[str, str] | None = None,
    estado_auto: str | None = None,
    en_calculo: bool | None = None,
) -> None:
    state = _level_state(dimension, level_id)
    if respuesta is not None:
        state["respuesta"] = respuesta
    if respuestas_preguntas is not None:
        state["respuestas_preguntas"] = respuestas_preguntas
    if evidencia is not None:
        state["evidencia"] = evidencia
    if evidencias_preguntas is not None:
        state["evidencias_preguntas"] = evidencias_preguntas
    if estado_auto is not None:
        state["estado_auto"] = estado_auto
    if en_calculo is not None:
        state["en_calculo"] = en_calculo
        if _EDIT_MODE_KEY not in st.session_state:
            st.session_state[_EDIT_MODE_KEY] = {}
        if dimension not in st.session_state[_EDIT_MODE_KEY]:
            st.session_state[_EDIT_MODE_KEY][dimension] = {}
        st.session_state[_EDIT_MODE_KEY][dimension][level_id] = not en_calculo
    if state.get("marcado_revision"):
        state["estado"] = "Revisión requerida"
    else:
        state["estado"] = state.get("estado_auto", "Pendiente")
    st.session_state[_STATE_KEY][dimension][level_id] = state


def _set_revision_flag(dimension: str, level_id: int, value: bool) -> None:
    state = _level_state(dimension, level_id)
    state["marcado_revision"] = value
    if state["marcado_revision"]:
        state["estado"] = "Revisión requerida"
    else:
        state["estado"] = state.get("estado_auto", "Pendiente")
    st.session_state[_STATE_KEY][dimension][level_id] = state


def _toggle_revision(dimension: str, level_id: int) -> None:
    state = _level_state(dimension, level_id)
    nuevo_valor = not state.get("marcado_revision", False)
    _set_revision_flag(dimension, level_id, nuevo_valor)

def _restore_level_form_values(dimension: str, level_id: int) -> None:
    """Restaura los valores del formulario para un nivel específico.
    
    Esta función maneja la restauración del estado del formulario, asegurando que:
    1. Los valores se restauren solo cuando sea necesario 
    2. No se sobrescriban estados válidos de widgets
    3. Se mantenga la consistencia entre toggles y respuestas
    4. Las evidencias se restauren correctamente
    """
    niveles = LEVEL_DEFINITIONS.get(dimension, [])
    level_data = next((lvl for lvl in niveles if lvl.get("nivel") == level_id), None)
    if not level_data:
        return
        
    state = _level_state(dimension, level_id)
    preguntas = level_data.get("preguntas") or []
    selector_key = f"selector_{dimension}_{level_id}"

    if preguntas:
        # Estado de respuestas y evidencias
        state_resp = state.get("respuestas_preguntas", {})
        state_evid = state.get("evidencias_preguntas", {}) 
        evidencias_agregadas = []

        # Restaurar cada pregunta individual
        for idx, _ in enumerate(preguntas, start=1):
            clave = str(idx)
            
            # Keys para esta pregunta
            resp_key = f"resp_{dimension}_{level_id}_{idx}"
            toggle_key = f"toggle_{dimension}_{level_id}_{idx}"
            evid_key = f"evid_{dimension}_{level_id}_{idx}"
            
            # Restaurar respuesta solo si no existe
            resp_valor = state_resp.get(clave)
            if resp_key not in st.session_state:
                st.session_state[resp_key] = resp_valor if resp_valor in {"VERDADERO", "FALSO"} else "FALSO"
                
            # Sincronizar toggle con respuesta
            if toggle_key not in st.session_state:
                st.session_state[toggle_key] = st.session_state[resp_key] == "VERDADERO"
            
            # Restaurar evidencia
            evid_texto = state_evid.get(clave, "")
            if evid_texto:
                if evid_key not in st.session_state:
                    st.session_state[evid_key] = evid_texto
                evidencias_agregadas.append(evid_texto.strip())
                
        # Agregar evidencias concatenadas
        evid_join_key = f"evid_{dimension}_{level_id}"
        if evid_join_key not in st.session_state and evidencias_agregadas:
            st.session_state[evid_join_key] = " \n".join(evidencias_agregadas)
            
        # Actualizar progreso
        total_preguntas = len(preguntas)
        progreso = _ensure_question_progress(dimension, level_id, total_preguntas)
        
        for idx in range(1, total_preguntas + 1):
            clave = str(idx)
            resp = state_resp.get(clave)
            evid = state_evid.get(clave, "")
            completo = _question_is_complete(resp, evid)
            progreso["saved"][clave] = completo
            
        progreso["active"] = 0
        st.session_state[_QUESTION_PROGRESS_KEY][dimension][level_id] = progreso
        st.session_state[selector_key] = 0
        
    else:
        # Para niveles sin preguntas individuales
        answer_key = f"resp_{dimension}_{level_id}"
        evidencia_key = f"evid_{dimension}_{level_id}"
        
        valor = state.get("respuesta")
        if answer_key not in st.session_state:
            st.session_state[answer_key] = valor if valor in {"VERDADERO", "FALSO"} else "FALSO"
            
        evidencia_val = state.get("evidencia", "")
        if evidencia_key not in st.session_state:
            st.session_state[evidencia_key] = "" if evidencia_val is None else str(evidencia_val)
            
    # Asegurar que el selector exista
    if selector_key not in st.session_state:
        st.session_state[selector_key] = 0


def _enqueue_level_restore(dimension: str, level_id: int) -> None:
    queue = st.session_state.get(_PENDING_RESTORE_QUEUE_KEY)
    if not isinstance(queue, list):
        queue = []
    item = (dimension, level_id)
    if item not in queue:
        queue.append(item)
    st.session_state[_PENDING_RESTORE_QUEUE_KEY] = queue


def _process_pending_restores(dimension: str) -> None:
    queue = st.session_state.get(_PENDING_RESTORE_QUEUE_KEY)
    if not queue:
        return
    if not isinstance(queue, list):
        queue = [queue]
    remaining: list[tuple[str, int]] = []
    for pending_entry in queue:
        if (
            not isinstance(pending_entry, tuple)
            or len(pending_entry) != 2
            or not isinstance(pending_entry[0], str)
            or not isinstance(pending_entry[1], int)
        ):
            continue
        pending_dimension, pending_level = pending_entry
        if pending_dimension == dimension:
            _restore_level_form_values(pending_dimension, pending_level)
            _update_ready_flag(pending_dimension, pending_level)
        else:
            remaining.append((pending_dimension, pending_level))
    st.session_state[_PENDING_RESTORE_QUEUE_KEY] = remaining


def _sync_dimension_score(dimension: str) -> int:
    niveles = LEVEL_DEFINITIONS.get(dimension, [])
    if not niveles:
        st.session_state["irl_scores"][dimension] = 0
        return 0
    niveles_sorted = sorted(niveles, key=lambda lvl: lvl.get("nivel", 0))
    baseline = niveles_sorted[0].get("nivel", 0)
    highest = baseline - 1
    for level in niveles_sorted:
        nivel_actual = level.get("nivel", baseline)
        expected_next = highest + 1
        if nivel_actual != expected_next:
            break
        level_state = _level_state(dimension, nivel_actual)
        if level_state["respuesta"] == "VERDADERO" and level_state["en_calculo"]:
            highest = nivel_actual
        else:
            break
    approved_level = highest if highest >= baseline else 0
    st.session_state["irl_scores"][dimension] = approved_level
    return approved_level


def _sync_all_scores() -> None:
    for dimension in STEP_TABS:
        _sync_dimension_score(dimension)


def _update_level_states_from_responses() -> None:
    """Actualiza los estados de los niveles basándose en las respuestas cargadas."""
    _init_irl_state()
    
    for dimension in STEP_TABS:
        niveles = LEVEL_DEFINITIONS.get(dimension, [])
        for level in niveles:
            level_id = level.get("nivel")
            if not level_id:
                continue
            
            preguntas = level.get("preguntas", [])
            
            # Verificar si todas las preguntas tienen respuesta VERDADERO
            if preguntas:
                todas_verdadero = True
                alguna_evidencia = False
                
                for idx in range(1, len(preguntas) + 1):
                    resp_key = f"resp_{dimension}_{level_id}_{idx}"
                    evid_key = f"evid_{dimension}_{level_id}_{idx}"
                    
                    respuesta = st.session_state.get(resp_key, "FALSO")
                    evidencia = st.session_state.get(evid_key, "")
                    
                    if respuesta != "VERDADERO":
                        todas_verdadero = False
                        break
                    
                    if evidencia and evidencia.strip() != "":
                        alguna_evidencia = True
                
                # Actualizar estado del nivel
                level_state = _level_state(dimension, level_id)
                if todas_verdadero:
                    level_state["respuesta"] = "VERDADERO"
                    # IMPORTANTE: Marcar en_calculo=True aunque no haya evidencia
                    # Para permitir cálculo de niveles desde Excel
                    level_state["en_calculo"] = True
                    level_state["estado"] = "Completo"
                    level_state["estado_auto"] = "Completo"
                    # Nota: En el flujo manual se puede requerir evidencia, pero desde Excel aceptamos sin evidencia
                else:
                    level_state["respuesta"] = "FALSO"
                    level_state["en_calculo"] = False
                    level_state["estado"] = "Incompleto"
                    level_state["estado_auto"] = "Incompleto"
            else:
                # Nivel sin preguntas específicas
                resp_key = f"resp_{dimension}_{level_id}"
                evid_key = f"evid_{dimension}_{level_id}"
                
                respuesta = st.session_state.get(resp_key, "FALSO")
                evidencia = st.session_state.get(evid_key, "")
                
                level_state = _level_state(dimension, level_id)
                level_state["respuesta"] = respuesta
                
                if respuesta == "VERDADERO":
                    # IMPORTANTE: Marcar en_calculo=True aunque no haya evidencia
                    level_state["en_calculo"] = True
                    level_state["estado"] = "Completo"
                    level_state["estado_auto"] = "Completo"
                else:
                    level_state["en_calculo"] = False
                    level_state["estado"] = "Incompleto"
                    level_state["estado_auto"] = "Incompleto"


def _compute_dimension_counts(dimension: str) -> dict:
    niveles = st.session_state[_STATE_KEY][dimension]
    total = len(niveles)
    completados = sum(1 for data in niveles.values() if data.get("en_calculo"))
    revision = sum(1 for data in niveles.values() if data.get("marcado_revision"))
    pendientes = max(total - completados, 0)
    return {
        "total": total,
        "completed": completados,
        "pending": pendientes,
        "revision": revision,
    }

def _dimension_badge_class(status: str) -> str:
    return {
        "Completa": "complete",
        "Parcial": "partial",
        "Pendiente": "pending",
    }.get(status, "pending")

def _dimension_badge(counts: dict) -> str:
    if counts["completed"] == counts["total"] and counts["revision"] == 0:
        return "Completa"
    if counts["completed"] or counts["revision"]:
        return "Parcial"
    return "Pendiente"


def _dimension_badge_class(status: str) -> str:
    return {
        "Completa": "complete",
        "Parcial": "partial",
        "Pendiente": "pending",
    }.get(status, "pending")


def _status_class(status: str) -> str:
    return _STATUS_CLASS_MAP.get(status, "pending")


def _normalize_question_responses(level: dict, respuestas: dict[str, str | None]) -> dict[str, str | None]:
    preguntas = level.get("preguntas") or []
    normalizado: dict[str, str | None] = {}
    for idx, _ in enumerate(preguntas, start=1):
        clave = str(idx)
        valor = respuestas.get(clave)
        normalizado[clave] = valor if valor in {"VERDADERO", "FALSO"} else None
    return normalizado


def _aggregate_question_status(respuestas: dict[str, str | None]) -> str | None:
    if not respuestas:
        return None
    valores = list(respuestas.values())
    if any(valor is None for valor in valores):
        return None
    if all(valor == "VERDADERO" for valor in valores):
        return "VERDADERO"
    return "FALSO"


def _handle_manual_answer_change(*, answer_key: str, evidencia_key: str) -> None:
    answer = st.session_state.get(answer_key)
    if answer != "VERDADERO":
        st.session_state[evidencia_key] = ""


def _handle_question_evidence_change(
    *,
    dimension: str,
    level_id: int,
    idx: int,
    total_questions: int,
    pregunta_key: str,
    evidencia_key: str,
) -> None:
    """Marca la pregunta como pendiente cuando cambia la evidencia.

    Importante: no escribimos en st.session_state[evidencia_key] aquí para
    evitar el error de Streamlit de modificar un widget luego de instanciarlo.
    Solo dejamos registro de que la pregunta tiene cambios pendientes.
    """
    _mark_question_pending(dimension, level_id, idx, total_questions)


def _handle_question_toggle_change(
    *,
    dimension: str,
    level_id: int,
    idx: int,
    total_questions: int,
    pregunta_key: str,
    evidencia_key: str,
    toggle_key: str,
) -> None:
    """Maneja el cambio en el toggle de una pregunta.
    
    Esta función se encarga de:
    1. Sincronizar el estado del toggle con la respuesta
    2. Limpiar evidencia cuando corresponda
    3. Marcar la pregunta como pendiente de guardar
    """
    # Obtener estado actual del toggle
    toggle_state = bool(st.session_state.get(toggle_key))
    
    # Convertir a VERDADERO/FALSO manteniendo consistencia
    nuevo_valor = "VERDADERO" if toggle_state else "FALSO"
    st.session_state[pregunta_key] = nuevo_valor
    
    # Si se cambia a FALSO, limpiar evidencia
    if not toggle_state:
        st.session_state[evidencia_key] = ""
        
    # Marcar como pendiente para forzar guardar
    _mark_question_pending(dimension, level_id, idx, total_questions)


def _persist_question_progress(
    dimension: str,
    level_id: int,
    idx: int,
    answer: str | None,
    evidence: str | None,
) -> None:
    level_state = _level_state(dimension, level_id)
    respuestas = dict(level_state.get("respuestas_preguntas") or {})
    evidencias = dict(level_state.get("evidencias_preguntas") or {})
    clave = str(idx)
    if answer in {"VERDADERO", "FALSO"}:
        respuestas[clave] = answer
    else:
        respuestas[clave] = None
    if answer == "VERDADERO":
        evidencias[clave] = _clean_text(evidence)
    else:
        evidencias[clave] = ""
    level_state["respuestas_preguntas"] = respuestas
    level_state["evidencias_preguntas"] = evidencias
    st.session_state[_STATE_KEY][dimension][level_id] = level_state


def _validate_level(dimension: str, level_id: int) -> tuple[bool, list[str]]:
    """Valida que el nivel tenga respuestas válidas y evidencias cuando corresponda."""
    errores: list[str] = []
    niveles = LEVEL_DEFINITIONS.get(dimension, [])
    level_data = next((lvl for lvl in niveles if lvl.get("nivel") == level_id), None)
    if not level_data:
        return False, ["Nivel no encontrado"]

    preguntas = level_data.get("preguntas") or []
    if preguntas:
        for idx, _ in enumerate(preguntas, start=1):
            resp_key = f"resp_{dimension}_{level_id}_{idx}"
            evid_key = f"evid_{dimension}_{level_id}_{idx}"
            resp = st.session_state.get(resp_key)
            if resp not in {"VERDADERO", "FALSO"}:
                errores.append(f"Pregunta {idx}: selecciona VERDADERO o FALSO.")
                continue
            if resp == "VERDADERO" and not _is_evidence_valid(st.session_state.get(evid_key)):
                min_chars = STEP_CONFIG.get("min_evidence_chars", 0)
                errores.append(
                    f"Pregunta {idx}: cuando es VERDADERO debes agregar antecedentes (mínimo {min_chars} caracteres si aplica)."
                )
    else:
        answer_key = f"resp_{dimension}_{level_id}"
        evidencia_key = f"evid_{dimension}_{level_id}"
        resp = st.session_state.get(answer_key)
        if resp not in {"VERDADERO", "FALSO"}:
            errores.append("Selecciona VERDADERO o FALSO.")
        elif resp == "VERDADERO" and not _is_evidence_valid(st.session_state.get(evidencia_key)):
            min_chars = STEP_CONFIG.get("min_evidence_chars", 0)
            errores.append(
                f"Cuando es VERDADERO debes agregar antecedentes (mínimo {min_chars} caracteres si aplica)."
            )

    return len(errores) == 0, errores

def _save_level_answers(
    dimension: str,
    level_id: int
) -> tuple[bool, str | None]:
    """Guarda las respuestas de un nivel después de validar.
    
    Esta función:
    1. Valida el nivel completo
    2. Recopila respuestas y evidencias
    3. Actualiza el estado del nivel
    4. Maneja errores
    
    Args:
        dimension: Código de la dimensión
        level_id: ID del nivel
        
    Returns:
        Tupla con:
        - bool: Si el guardado fue exitoso
        - str: Mensaje de error si hay
    """
    # Primero validamos todo el nivel
    valid, errors = _validate_level(dimension, level_id)
    if not valid:
        error_msg = "\n".join(errors)
        return False, error_msg
        
    # Obtener datos del nivel
    niveles = LEVEL_DEFINITIONS.get(dimension, [])
    level_data = next((lvl for lvl in niveles if lvl.get("nivel") == level_id), None)
    if not level_data:
        return False, "Nivel no encontrado"
        
    preguntas = level_data.get("preguntas") or []
    state = _level_state(dimension, level_id)
    
    if preguntas:
        # Recopilar respuestas y evidencias
        respuestas: dict[str, str] = {}
        evidencias: dict[str, str] = {}
        
        for idx, _ in enumerate(preguntas, start=1):
            pregunta_key = f"resp_{dimension}_{level_id}_{idx}"
            evidencia_key = f"evid_{dimension}_{level_id}_{idx}"
            clave = str(idx)
            
            respuesta = st.session_state.get(pregunta_key)
            if respuesta not in {"VERDADERO", "FALSO"}:
                respuesta = "FALSO"
            respuestas[clave] = respuesta
            
            evidencia = st.session_state.get(evidencia_key, "").strip()
            evidencias[clave] = evidencia if respuesta == "VERDADERO" else ""
            
        # Determinar respuesta agregada
        respuesta = _aggregate_question_status(respuestas) or "FALSO"
        evidencia = " \n".join(texto for texto in evidencias.values() if texto).strip()
        
    else:
        # Nivel sin preguntas
        answer_key = f"resp_{dimension}_{level_id}"
        evidencia_key = f"evid_{dimension}_{level_id}"
        
        respuesta = st.session_state.get(answer_key)
        if respuesta not in {"VERDADERO", "FALSO"}:
            respuesta = "FALSO"
            
        evidencia = st.session_state.get(evidencia_key, "").strip()
        if respuesta != "VERDADERO":
            evidencia = ""
            
        respuestas = {}
        evidencias = {}
        
    # Actualizar estado
    _set_level_state(
        dimension,
        level_id,
        respuesta=respuesta,
        respuestas_preguntas=respuestas,
        evidencia=evidencia,
        evidencias_preguntas=evidencias,
        estado_auto="Respondido (en cálculo)",
        en_calculo=True
    )
    
    return True, None

def _handle_level_submission(
    dimension: str,
    level_id: int,
    respuestas_preguntas: dict[str, str | None],
    evidencia: str,
    *,
    evidencias_preguntas: dict[str, str] | None = None,
    respuesta_manual: str | None = None,
) -> tuple[bool, str | None, str | None]:
    """Punto de entrada para guardar un nivel.
    
    Esta función delega la validación y guardado a _save_level_answers,
    pero mantiene la interfaz existente para compatibilidad.
    """
    success, error = _save_level_answers(dimension, level_id)
    return success, error, None


# Nota: Implementación única de _render_level_question_flow definida más arriba.


def _render_dimension_tab(dimension: str) -> None:
    """Renderiza una pestaña de dimensión completa."""
    # Inicializar estado
    _init_irl_state()
    _process_pending_restores(dimension)
    
    # Obtener definición de niveles
    levels = LEVEL_DEFINITIONS.get(dimension, [])
    # Calcular estadísticas
    counts = _compute_dimension_counts(dimension)
    
    # Mostrar banner si hay mensaje
    banner_msg = st.session_state.get(_BANNER_KEY, {}).get(dimension)
    if banner_msg:
        st.info(banner_msg)

    st.markdown(IRL_IMPORTANT_HTML, unsafe_allow_html=True)

    progreso = counts["completed"] / counts["total"] if counts["total"] else 0
    st.markdown(
        f"**{counts['completed']} de {counts['total']} niveles respondidos**"
    )
    st.progress(progreso)

    for lvl_index, level in enumerate(levels):
        level_id = level["nivel"]
        state = _level_state(dimension, level_id)
        status = state.get("estado", "Pendiente")
        status_class = _status_class(status)
        card_classes = ["level-card", f"level-card--{status_class}"]
        if state.get("en_calculo"):
            card_classes.append("level-card--answered")
        if st.session_state[_ERROR_KEY][dimension].get(level_id):
            card_classes.append("level-card--error")
        edit_mode = st.session_state[_EDIT_MODE_KEY][dimension].get(
            level_id,
            not state.get("en_calculo"),
        )
        locked = bool(state.get("en_calculo")) and not edit_mode

        restore_flags = st.session_state[_RESTORE_ON_EDIT_KEY][dimension]
        if not edit_mode or locked:
            restore_flags[level_id] = False
        elif state.get("en_calculo"):
            if not restore_flags.get(level_id):
                _restore_level_form_values(dimension, level_id)
                restore_flags[level_id] = True
        else:
            restore_flags[level_id] = False

        if locked:
            card_classes.append("level-card--locked")
        elif edit_mode:
            card_classes.append("level-card--editing")

        st.markdown(
            f"<div class='{' '.join(card_classes)}' id='{dimension}-{level_id}'>",
            unsafe_allow_html=True,
        )

        expander_label = f"Nivel {level_id} · {level['descripcion']}"
        # Control expanders explicitly per-level to avoid racey JS-based close behaviour.
        expander_open_key = f"expander_open_{dimension}_{level_id}"
        # If there's an error for this level, force the expander open. Otherwise respect stored flag.
        expanded = bool(st.session_state[_ERROR_KEY][dimension].get(level_id)) or bool(
            st.session_state.get(expander_open_key, False)
        )
        with st.expander(
            expander_label,
            expanded=expanded,
        ):
            preguntas = level.get("preguntas") or []
            answer_key = f"resp_{dimension}_{level_id}"
            evidencia_key = f"evid_{dimension}_{level_id}"
            if evidencia_key not in st.session_state:
                evidencia_val = state.get("evidencia", "")
                st.session_state[evidencia_key] = "" if evidencia_val is None else str(evidencia_val)

            respuestas_dict: dict[str, str | None] = {}
            evidencias_dict_envio: dict[str, str] | None = None
            evidencia_texto = st.session_state.get(evidencia_key, "")
            respuesta_manual: str | None = None
            ready_to_save = False

            show_cancel = bool(state.get("en_calculo")) and edit_mode and not locked
            editar_label = "Cancelar" if show_cancel else "Editar"
            editar_disabled = False
            if not state.get("en_calculo") and edit_mode:
                editar_disabled = True

            if preguntas:
                (
                    respuestas_dict,
                    evidencias_dict_envio,
                    evidencia_texto,
                    ready_to_save,
                ) = _render_level_question_flow(
                    dimension,
                    level_id,
                    preguntas,
                    level.get("descripcion", ""),
                    locked=locked,
                )
            else:
                current_answer = state.get("respuesta")
                current_option = (
                    current_answer if current_answer in {"VERDADERO", "FALSO"} else "FALSO"
                )
                if answer_key not in st.session_state:
                    st.session_state[answer_key] = current_option

                st.radio(
                    "Responder",
                    options=["VERDADERO", "FALSO"],
                    key=answer_key,
                    horizontal=True,
                    disabled=locked,
                    on_change=_handle_manual_answer_change,
                    kwargs={
                        "answer_key": answer_key,
                        "evidencia_key": evidencia_key,
                    },
                )

                respuesta_manual = st.session_state.get(answer_key)
                evidencia_texto = st.text_area(
                    "Antecedentes de verificación",
                    key=evidencia_key,
                    placeholder="Describe brevemente los antecedentes que respaldan esta afirmación…",
                    height=110,
                    max_chars=STEP_CONFIG["max_char_limit"],
                    disabled=locked or respuesta_manual != "VERDADERO",
                )

                if respuesta_manual == "VERDADERO":
                    contador = len(_clean_text(evidencia_texto))
                    contador_html = (
                        f"<div class='stepper-form__counter{' stepper-form__counter--alert' if contador > STEP_CONFIG['soft_char_limit'] else ''}'>"
                        f"{contador}/{STEP_CONFIG['soft_char_limit']}"
                        "</div>"
                    )
                    st.markdown(contador_html, unsafe_allow_html=True)
                else:
                    st.caption("Disponible solo si seleccionas VERDADERO.")

                ready_to_save = respuesta_manual in {"VERDADERO", "FALSO"}
                if ready_to_save and respuesta_manual == "VERDADERO":
                    ready_to_save = _is_evidence_valid(evidencia_texto)

                evidencias_dict_envio = None
                st.session_state[_READY_KEY][dimension][level_id] = ready_to_save

            error_msg = st.session_state[_ERROR_KEY][dimension].get(level_id)
            if error_msg:
                st.error(error_msg)

            guardar = st.button(
                "Guardar y continuar con el siguiente nivel",
                type="primary",
                disabled=locked or not ready_to_save,
                key=f"btn_guardar_{dimension}_{level_id}",
                use_container_width=True,
            )

            if guardar:
                success, error_message, banner = _handle_level_submission(
                    dimension,
                    level_id,
                    respuestas_dict,
                    evidencia_texto,
                    evidencias_preguntas=evidencias_dict_envio,
                    respuesta_manual=respuesta_manual,
                )
                st.session_state[_BANNER_KEY][dimension] = banner
                if error_message:
                    st.session_state[_ERROR_KEY][dimension][level_id] = error_message
                else:
                    st.session_state[_ERROR_KEY][dimension][level_id] = None
                    _sync_dimension_score(dimension)
                    _set_revision_flag(dimension, level_id, False)
                    st.session_state[_EDIT_MODE_KEY][dimension][level_id] = False
                    # Close current expander and open the next one (if any) deterministically
                    st.session_state[_CLOSE_EXPANDER_KEY] = (dimension, level_id)
                    # close current
                    st.session_state[f"expander_open_{dimension}_{level_id}"] = False
                    # open next level if exists
                    if lvl_index + 1 < len(levels):
                        next_level_id = levels[lvl_index + 1]["nivel"]
                        st.session_state[f"expander_open_{dimension}_{next_level_id}"] = True

                    irl_level_flow.save_level("Nivel guardado")
                    st.toast("Guardado")
                    # Recalcular puntaje global y cachearlo para evitar cálculos en cada rerun
                    try:
                        df_all = _collect_dimension_responses()
                        if not df_all.empty:
                            st.session_state["irl_last_puntaje"] = trl.calcular_trl(
                                df_all[["dimension", "nivel", "evidencia"]]
                            )
                        else:
                            st.session_state["irl_last_puntaje"] = None
                    except Exception:
                        st.session_state["irl_last_puntaje"] = None
                    _rerun_app()
        if st.session_state.get(_CLOSE_EXPANDER_KEY) == (dimension, level_id):
            st.session_state[_CLOSE_EXPANDER_KEY] = None
            components.html(
                f"""
                <script>
                const container = window.parent.document.getElementById('{dimension}-{level_id}');
                if (container) {{
                    const details = container.querySelector("div[data-testid='stExpander'] details");
                    if (details) {{ details.open = false; }}
                }}
                </script>
                """,
                height=0,
            )

        st.markdown("</div>", unsafe_allow_html=True)

def _collect_dimension_responses() -> pd.DataFrame:
    _init_irl_state()
    dimensiones_ids = trl.ids_dimensiones()
    etiquetas = dict(zip(dimensiones_ids, trl.labels_dimensiones()))
    registros: list[dict] = []

    for dimension in dimensiones_ids:
        niveles = LEVEL_DEFINITIONS.get(dimension, [])
        evidencias: list[str] = []
        if not niveles:
            st.session_state["irl_scores"][dimension] = 0
            registros.append(
                {
                    "dimension": dimension,
                    "etiqueta": etiquetas.get(dimension, dimension),
                    "nivel": None,
                    "evidencia": "",
                }
            )
            continue
        niveles_sorted = sorted(niveles, key=lambda lvl: lvl.get("nivel", 0))
        baseline = niveles_sorted[0].get("nivel", 0)
        highest = baseline - 1
        for level in niveles_sorted:
            nivel_actual = level.get("nivel", baseline)
            expected_next = highest + 1
            if nivel_actual != expected_next:
                break
            data = _level_state(dimension, nivel_actual)
            if data.get("respuesta") == "VERDADERO" and data.get("en_calculo"):
                highest = nivel_actual
                evidencia_txt = (data.get("evidencia") or "").strip()
                if evidencia_txt:
                    evidencias.append(evidencia_txt)
            else:
                break
        approved_level = highest if highest >= baseline else 0
        st.session_state["irl_scores"][dimension] = approved_level
        registros.append(
            {
                "dimension": dimension,
                "etiqueta": etiquetas.get(dimension, dimension),
                "nivel": approved_level if approved_level else None,
                "evidencia": " · ".join(evidencias),
            }
        )

    return pd.DataFrame(registros)



def _level_has_response(state: dict | None) -> bool:
    if not state:
        return False

    estado_auto = state.get("estado_auto")
    if estado_auto and estado_auto != "Pendiente":
        return True

    if state.get("en_calculo"):
        return True

    estado = state.get("estado")
    if estado and estado != "Pendiente":
        return True

    if state.get("marcado_revision"):
        return True

    return False


def _format_answer_display(valor: str | None, state: dict | None) -> str:
    if valor == "VERDADERO":
        return "Verdadero"

    if valor == "FALSO":
        return "Falso" if _level_has_response(state) else "No respondido"

    return "No respondido"


def _collect_dimension_details() -> dict[str, dict[str, Any]]:
    _init_irl_state()
    dimensiones_ids = trl.ids_dimensiones()
    etiquetas = dict(zip(dimensiones_ids, trl.labels_dimensiones()))
    detalles: dict[str, dict[str, Any]] = {}

    for dimension in dimensiones_ids:
        niveles = LEVEL_DEFINITIONS.get(dimension, [])
        filas: list[dict[str, Any]] = []
        for level in niveles:
            nivel_id = level.get("nivel")
            state = _level_state(dimension, nivel_id)
            estado_nivel = state.get("estado", "Pendiente")
            preguntas = level.get("preguntas") or []
            if preguntas:
                respuestas = state.get("respuestas_preguntas") or {}
                evidencias_preguntas = state.get("evidencias_preguntas") or {}
                for idx, pregunta in enumerate(preguntas, start=1):
                    idx_str = str(idx)
                    filas.append(
                        {
                            "Nivel": nivel_id,
                            "Descripción del nivel": level.get("descripcion", ""),
                            "Pregunta": pregunta,
                            "Respuesta": _format_answer_display(
                                respuestas.get(idx_str), state
                            ),
                            "Antecedentes de verificación": evidencias_preguntas.get(idx_str) or "—",
                            "Estado del nivel": estado_nivel,
                        }
                    )
            else:
                filas.append(
                    {
                        "Nivel": nivel_id,
                        "Descripción del nivel": level.get("descripcion", ""),
                        "Pregunta": "—",
                        "Respuesta": _format_answer_display(state.get("respuesta"), state),
                        "Antecedentes de verificación": state.get("evidencia") or "—",
                        "Estado del nivel": estado_nivel,
                    }
                )

        detalles[dimension] = {
            "label": etiquetas.get(dimension, dimension),
            "rows": filas,
        }

    return detalles


st.set_page_config(page_title="Fase 1 - Evaluación IRL", page_icon="🌲", layout="wide")
_safe_load_theme()

st.markdown(
    """
<style>
.page-intro {
    display: grid;
    grid-template-columns: minmax(0, 1.6fr) minmax(0, 1fr);
    gap: 1.0rem;
    padding: 1.0rem 1.2rem;
    border-radius: 10px;
    background: linear-gradient(145deg, rgba(18, 48, 29, 0.9), rgba(111, 75, 44, 0.82));
    color: #fdf9f2;
    box-shadow: 0 8px 18px rgba(12, 32, 20, 0.2);
    margin-bottom: 1.0rem;
}

.page-intro h1 {
    font-size: 2.2rem;
    margin-bottom: 1rem;
    color: #fffdf8;
}

.page-intro p {
    font-size: 1.02rem;
    line-height: 1.6;
    color: rgba(253, 249, 242, 0.86);
}

.page-intro__aside {
    display: flex;
    flex-direction: column;
    gap: 1rem;
}

.page-intro__aside .intro-stat {
    background: rgba(255, 255, 255, 0.14);
    border-radius: 20px;
    padding: 1.1rem 1.3rem;
    box-shadow: inset 0 0 0 1px rgba(255, 255, 255, 0.12);
}

.page-intro__aside .intro-stat strong {
    display: block;
    text-transform: uppercase;
    letter-spacing: 0.6px;
    font-size: 0.9rem;
    margin-bottom: 0.35rem;
    color: #fefcf9;
}

.page-intro__aside .intro-stat p {
    margin: 0;
    color: rgba(253, 249, 242, 0.86);
    font-size: 0.96rem;
    line-height: 1.5;
}

.back-band {
    display: flex;
    justify-content: flex-end;
    margin-bottom: 1.6rem;
}

.metric-ribbon {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(190px, 1fr));
    gap: 1.1rem;
    margin: 1.4rem 0 2.1rem;
}

.metric-ribbon__item {
    background: #ffffff;
    border-radius: 20px;
    padding: 1.3rem 1.4rem;
    border: 1px solid rgba(var(--shadow-color), 0.12);
    box-shadow: 0 20px 42px rgba(var(--shadow-color), 0.16);
    position: relative;
    overflow: hidden;
}

.metric-ribbon__item:after {
    content: "";
    position: absolute;
    width: 120px;
    height: 120px;
    border-radius: 50%;
    background: rgba(37, 87, 52, 0.12);
    top: -40px;
    right: -50px;
}

.metric-ribbon__value {
    font-size: 2.1rem;
    font-weight: 700;
    color: var(--forest-700);
    position: relative;
}

.metric-ribbon__label {
    display: block;
    margin-top: 0.4rem;
    font-size: 0.78rem;
    font-weight: 600;
    letter-spacing: 0.55px;
    text-transform: uppercase;
    color: var(--text-500);
}

.section-shell {
    background: transparent;
    border-radius: 8px;
    padding: 0.4rem 0.4rem;
    border: none;
    box-shadow: none;
    margin-bottom: 0.8rem;
}

.section-shell--split {
    padding: 0.4rem 0.4rem 0.6rem;
}

.section-shell h3, .section-shell h4 {
    margin-top: 0;
}

.threshold-band {
    display: flex;
    flex-wrap: wrap;
    gap: 0.6rem;
    margin: 0.6rem 0 1rem;
}

.threshold-chip {
    display: inline-flex;
    align-items: center;
    gap: 0.5rem;
    padding: 0.45rem 1rem;
    border-radius: 16px;
    background: rgba(var(--forest-500), 0.16);
    color: var(--text-700);
    font-weight: 600;
    border: 1px solid rgba(var(--forest-500), 0.22);
}

.threshold-chip strong {
    font-size: 1rem;
    color: var(--forest-700);
}

.selection-card {
    position: relative;
    padding: 1.2rem 1.5rem;
    border-radius: 14px;
    background: linear-gradient(135deg, #ffffff 0%, #f8fafb 100%);
    border: 2px solid #1b5e20;
    box-shadow: 0 4px 16px rgba(27, 94, 32, 0.12), 0 2px 6px rgba(0,0,0,0.06);
    overflow: hidden;
    transition: transform 180ms ease, box-shadow 180ms ease;
}

.selection-card:hover {
    transform: translateY(-2px);
    box-shadow: 0 6px 20px rgba(27, 94, 32, 0.16), 0 3px 8px rgba(0,0,0,0.08);
}

.selection-card::before {
    content: '';
    position: absolute;
    top: 0;
    left: 0;
    width: 100%;
    height: 5px;
    background: linear-gradient(90deg, #1b5e20 0%, #43a047 50%, #1b5e20 100%);
}

.selection-card::after {
    content: '';
    position: absolute;
    top: -50%;
    right: -10%;
    width: 280px;
    height: 280px;
    background: radial-gradient(circle, rgba(27,94,32,0.04) 0%, transparent 70%);
    border-radius: 50%;
    pointer-events: none;
}

.selection-card__badge {
    display: inline-flex;
    align-items: center;
    gap: 0.5rem;
    padding: 0.35rem 0.85rem;
    border-radius: 999px;
    background: linear-gradient(135deg, #1b5e20 0%, #2e7d32 100%);
    color: #ffffff;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    font-size: 0.72rem;
    font-weight: 700;
    box-shadow: 0 2px 8px rgba(27, 94, 32, 0.25);
    position: relative;
    z-index: 2;
}

.selection-card__badge::before {
    content: '✓';
    font-size: 0.85rem;
    font-weight: 900;
}

.selection-card__title {
    margin: 0.8rem 0 0.4rem;
    font-size: 1.4rem;
    font-weight: 700;
    color: #1b5e20;
    line-height: 1.3;
    position: relative;
    z-index: 2;
}

.selection-card__subtitle {
    margin: 0 0 1rem;
    color: #2e7d32;
    font-size: 1rem;
    font-weight: 500;
    position: relative;
    z-index: 2;
}

.selection-card__meta {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
    gap: 0.7rem;
    margin-top: 1rem;
    position: relative;
    z-index: 2;
}

.selection-card__meta-item {
    padding: 0.7rem 0.9rem;
    border-radius: 10px;
    background: rgba(255, 255, 255, 0.75);
    border: 1px solid rgba(27, 94, 32, 0.15);
    box-shadow: 0 1px 4px rgba(0,0,0,0.04);
    backdrop-filter: blur(4px);
    transition: background 150ms ease, border-color 150ms ease;
}

.selection-card__meta-item:hover {
    background: rgba(255, 255, 255, 0.95);
    border-color: rgba(27, 94, 32, 0.25);
}

.selection-card__meta-label {
    display: flex;
    align-items: center;
    gap: 0.3rem;
    text-transform: uppercase;
    font-size: 0.7rem;
    letter-spacing: 0.5px;
    font-weight: 700;
    color: #5a7d5e;
    margin-bottom: 0.3rem;
}

.selection-card__meta-label::before {
    content: '▪';
    color: #43a047;
    font-size: 0.9rem;
}

.selection-card__meta-value {
    display: block;
    font-size: 1rem;
    font-weight: 600;
    color: #1b3c1f;
    line-height: 1.3;
}

.history-caption {
    color: var(--text-500);
    margin-bottom: 0.8rem;
}

/* Panel compacto para el detalle de niveles */
.details-panel {
    padding: 0.4rem 0.5rem;
}

.details-panel h3,
.details-panel h4 {
    margin: 0.2rem 0 0.4rem;
}

.details-panel p,
.details-panel .stMarkdown p {
    margin: 0.2rem 0;
}

.details-panel strong,
.details-panel .stMarkdown strong {
    font-weight: 600;
}

.details-panel div[data-testid="stTabs"] {
    margin-top: 0.2rem;
}

@media (max-width: 992px) {
    .page-intro {
        grid-template-columns: 1fr;
    }

    .back-band {
        justify-content: center;
    }
}

div[data-testid="stExpander"] {
    margin-bottom: 0.2rem;
}

div[data-testid="stExpander"] > details {
    border-radius: 6px;
    border: 1px solid rgba(var(--shadow-color), 0.15);
    background: #ffffff;
    box-shadow: 0 2px 6px rgba(var(--shadow-color), 0.08);
    overflow: hidden;
}

div[data-testid="stExpander"] > details > summary {
    font-weight: 700;
    font-size: 0.9rem;
    color: var(--forest-700);
    padding: 0.5rem 0.7rem;
    list-style: none;
    position: relative;
}

div[data-testid="stExpander"] > details > summary::before {
    content: "➕";
    margin-right: 0.6rem;
    color: var(--forest-600);
    font-size: 1rem;
}

div[data-testid="stExpander"] > details[open] > summary::before {
    content: "➖";
}

div[data-testid="stExpander"] > details[open] > summary {
    background: #f5f7f9;
    color: var(--text-700);
}

div[data-testid="stExpander"] > details > div[data-testid="stExpanderContent"] {
    padding: 0.5rem 0.7rem 0.7rem;
    background: #ffffff;
    border-top: 1px solid rgba(var(--shadow-color), 0.15);
}

.irl-bubbles {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
    gap: 0.85rem;
    margin: 1rem 0 1.4rem;
}

.irl-bubble {
    border-radius: 18px;
    padding: 0.85rem 1rem;
    background: rgba(var(--forest-100), 0.72);
    border: 1px solid rgba(var(--forest-500), 0.25);
    box-shadow: inset 0 0 0 1px rgba(255, 255, 255, 0.6);
    display: flex;
    flex-direction: column;
    gap: 0.25rem;
}

.irl-bubble__label {
    font-size: 0.95rem;
    font-weight: 600;
    color: var(--forest-800);
}

.irl-bubble__badge {
    font-size: 0.82rem;
    text-transform: uppercase;
    letter-spacing: 0.55px;
}

.irl-bubble small {
    color: var(--text-500);
    font-size: 0.75rem;
}

.irl-bubble--complete {
    background: rgba(46, 142, 86, 0.18);
    border-color: rgba(46, 142, 86, 0.45);
}

.irl-bubble--partial {
    background: rgba(234, 185, 89, 0.22);
    border-color: rgba(234, 185, 89, 0.45);
}

.irl-bubble--pending {
    background: rgba(180, 196, 210, 0.25);
    border-color: rgba(143, 162, 180, 0.42);
}

.irl-important {
    margin: 0.6rem 0 1.1rem;
    padding: 0.85rem 1rem;
    border-radius: 16px;
    background: linear-gradient(135deg, rgba(56, 116, 209, 0.16), rgba(21, 118, 78, 0.14));
    border: 1px solid rgba(56, 116, 209, 0.25);
    color: rgba(26, 44, 84, 0.92);
    font-size: 0.9rem;
    line-height: 1.5;
    box-shadow: 0 12px 24px rgba(var(--shadow-color), 0.12);
}

.irl-important strong {
    text-transform: uppercase;
    letter-spacing: 0.6px;
}

.irl-important a {
    color: rgba(12, 74, 50, 0.95);
    font-weight: 700;
    text-decoration: none;
    border-bottom: 1px solid rgba(12, 74, 50, 0.4);
}

.irl-important a:hover {
    color: rgba(12, 74, 50, 0.8);
}

.irl-important__hint {
    display: block;
    margin-top: 0.4rem;
    font-size: 0.82rem;
    color: rgba(26, 44, 84, 0.78);
}

.level-card {
    border-radius: 8px;
    border: 1px solid rgba(var(--shadow-color), 0.1);
    background: #ffffff;
    box-shadow: none;
    margin-bottom: 0.25rem;
    transition: border-color 0.15s ease, box-shadow 0.15s ease;
}

.level-card:hover {
    box-shadow: none;
}

.level-card > div[data-testid="stExpander"] > details {
    border: none;
    background: transparent;
}

.level-card > div[data-testid="stExpander"] > details > summary {
    font-size: 0.9rem;
    font-weight: 700;
    color: var(--text-800);
    padding: 0.5rem 0.7rem;
    list-style: none;
    cursor: pointer;
}

.level-card > div[data-testid="stExpander"] > details > summary::-webkit-details-marker {
    display: none;
}

.level-card > div[data-testid="stExpander"] div[data-testid="stExpanderContent"] {
    padding: 0 0.7rem 0.6rem;
    background: #ffffff;
    border-top: 1px solid rgba(var(--shadow-color), 0.15);
}

.level-card--answered {
    border-color: rgba(58, 181, 112, 0.35);
    box-shadow: none;
    background: #ffffff;
}

.level-card--editing {
    border-color: rgba(21, 118, 78, 0.35);
    box-shadow: none;
}

.level-card--editing > div[data-testid="stExpander"] > details > summary {
    color: var(--text-800);
}

.level-card--locked {
    background: #f7f9fb;
    border-color: rgba(135, 145, 163, 0.35);
    box-shadow: none;
}

.level-card--locked > div[data-testid="stExpander"] > details > summary {
    color: rgba(46, 59, 79, 0.88);
    text-shadow: none;
}

.level-card--locked .level-card__intro {
    color: rgba(48, 61, 80, 0.8);
}

.level-card--complete {
    border-color: rgba(30, 78, 155, 0.35);
}

.level-card--complete > div[data-testid="stExpander"] > details > summary {
    color: var(--text-800);
}

.level-card--answered > div[data-testid="stExpander"] > details > summary {
    background: #f5f7f9;
    color: var(--text-800);
    text-shadow: none;
}

.level-card--answered > div[data-testid="stExpander"] > details[open] > summary {
    background: #eef2f5;
    color: var(--text-800);
}

.level-card--answered > div[data-testid="stExpander"] > details > summary::before {
    color: var(--forest-600);
}

.level-card--answered > div[data-testid="stExpander"] > details > summary::after {
    display: none;
}

.level-card--pending {
    border-color: rgba(143, 162, 180, 0.25);
    background: #ffffff;
}

.level-card--attention {
    border-color: rgba(224, 156, 70, 0.35);
    background: #ffffff;
}

.level-card--review {
    border-color: rgba(156, 112, 230, 0.35);
    background: #ffffff;
}

.level-card--error {
    border-color: rgba(206, 104, 86, 0.45);
    box-shadow: none;
    background: #ffffff;
}

.level-card__intro {
    font-size: 0.92rem;
    color: var(--text-600);
    margin-bottom: 0.75rem;
    line-height: 1.45;
}

.question-block {
    border: 1px solid rgba(var(--shadow-color), 0.1);
    border-radius: 6px;
    padding: 0.5rem 0.6rem 0.5rem;
    margin-bottom: 0.25rem;
    background: #ffffff;
    box-shadow: none;
    transition: border-color 0.15s ease, background 0.15s ease;
}

.question-block--true {
    background: #ffffff;
    border-color: rgba(21, 118, 78, 0.35);
}

.question-block--pending {
    background: #ffffff;
    border-color: rgba(134, 149, 170, 0.3);
}

.question-block--false {
    background: #ffffff;
    border-color: rgba(120, 135, 155, 0.28);
}

.question-block--saved {
    box-shadow: none;
}

.question-block--locked {
    background: #f7f9fb;
    box-shadow: none;
    opacity: 0.85;
}

.question-block--locked .question-block__chip {
    filter: grayscale(0.4);
    opacity: 0.85;
}

.question-block__header {
    display: flex;
    gap: 0.6rem;
    align-items: flex-start;
    font-size: 0.92rem;
    font-weight: 600;
    color: var(--text-700);
}

.question-block__body {
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 0.75rem;
    width: 100%;
}

.question-block__badge {
    min-width: 1.5rem;
    height: 1.5rem;
    border-radius: 999px;
    background: #eef2f5;
    color: var(--text-700);
    display: flex;
    align-items: center;
    justify-content: center;
    font-weight: 700;
    font-size: 0.78rem;
}

.question-block__text {
    flex: 1;
    line-height: 1.35;
}

.question-block__chip {
    border-radius: 999px;
    padding: 0.12rem 0.6rem;
    font-size: 0.75rem;
    font-weight: 700;
    letter-spacing: 0.4px;
    text-transform: uppercase;
}

.question-block__chip--true {
    background: #f3f5f7;
    color: var(--text-700);
    border: 1px solid rgba(var(--shadow-color), 0.08);
}

.question-block__chip--false {
    background: #f3f5f7;
    color: var(--text-700);
    border: 1px solid rgba(var(--shadow-color), 0.08);
}

.question-block__chip--pending {
    background: #f3f5f7;
    color: var(--text-700);
    border: 1px solid rgba(var(--shadow-color), 0.08);
}

.question-block__chip--draft {
    box-shadow: inset 0 0 0 1px rgba(58, 76, 102, 0.22);
}

.question-block__counter {
    text-align: right;
    margin-top: 0.25rem;
    font-size: 0.75rem;
    color: rgba(var(--shadow-color), 0.65);
}

.question-block__counter--alert {
    color: rgba(184, 108, 54, 0.85);
    font-weight: 600;
}

.question-stepper {
    display: flex;
    flex-wrap: wrap;
    gap: 0.35rem;
    margin-bottom: 0.5rem;
}

.question-stepper__item {
    min-width: 2.2rem;
    padding: 0.3rem 0.6rem;
    border-radius: 10px;
    background: #ffffff;
    color: var(--text-700);
    font-weight: 600;
    font-size: 0.85rem;
    text-align: center;
    border: 1px solid rgba(var(--shadow-color), 0.08);
    transition: background 0.15s ease, color 0.15s ease, border-color 0.15s ease;
}

.question-stepper__item.is-done {
    background: #f5f7f9;
    color: var(--text-700);
    box-shadow: none;
}

.question-stepper__item.is-active {
    background: #eef2f5;
    color: var(--text-800);
    box-shadow: none;
    border-color: rgba(var(--shadow-color), 0.12);
}

.question-stepper__item.is-active.is-done {
    background: #eef2f5;
}

.question-actions {
    margin-top: 0.6rem;
}

.question-actions > div[data-testid="column"] {
    display: flex;
    flex-direction: column;
}

.question-actions > div[data-testid="column"] > div {
    width: 100%;
}

.question-action {
    width: 100%;
}

.question-action > div[data-testid="stButton"] > button {
    width: 100%;
    border-radius: 12px;
    font-weight: 700;
    transition: transform 0.2s ease, box-shadow 0.2s ease, background 0.2s ease;
}

.question-action--next > div[data-testid="stButton"] > button,
.question-action--save > div[data-testid="stButton"] > button {
    background: linear-gradient(135deg, #1e9d6c, #15754e);
    color: #ffffff;
    border: 1px solid rgba(17, 94, 63, 0.85);
    box-shadow: 0 12px 22px rgba(21, 117, 78, 0.24);
}

.question-action--next > div[data-testid="stButton"] > button:hover:enabled,
.question-action--save > div[data-testid="stButton"] > button:hover:enabled {
    background: linear-gradient(135deg, #25b27c, #1b8a5d);
    box-shadow: 0 16px 28px rgba(21, 117, 78, 0.28);
    transform: translateY(-1px);
}

.question-action--next > div[data-testid="stButton"] > button:disabled,
.question-action--save > div[data-testid="stButton"] > button:disabled {
    background: linear-gradient(135deg, #e5e7eb, #d1d5db);
    color: #1f2937;
    border: 1px solid #9ca3af;
    box-shadow: none;
    cursor: not-allowed;
    opacity: 1;
}

.question-action--prev > div[data-testid="stButton"] > button {
    background: rgba(31, 55, 91, 0.08);
    color: rgba(28, 53, 88, 0.85);
    border: 1px solid rgba(28, 53, 88, 0.14);
    box-shadow: none;
}

.question-action--prev > div[data-testid="stButton"] > button:hover:enabled {
    background: rgba(31, 55, 91, 0.12);
    color: rgba(28, 53, 88, 0.95);
}

.question-action--prev > div[data-testid="stButton"] > button:disabled {
    opacity: 0.55;
    cursor: not-allowed;
}

.question-action--single {
    margin-top: 0.6rem;
}

.question-toggle {
    display: flex;
    flex-direction: column;
    align-items: flex-end;
    gap: 0.35rem;
}

.question-toggle > div[data-testid="stToggle"] {
    width: 100%;
    display: flex;
    justify-content: flex-end;
}

.question-toggle > div[data-testid="stToggle"] label {
    transform: none;
}

.question-toggle__state {
    font-size: 0.82rem;
    font-weight: 700;
    letter-spacing: 0.4px;
}

.question-toggle__state--true {
    color: rgba(17, 94, 63, 0.95);
}

.question-toggle__state--false {
    color: rgba(130, 32, 32, 0.92);
}

.level-card--locked .question-block__counter,
.level-card--locked .stepper-form__counter {
    opacity: 0.65;
}

.level-card--locked .stTextArea textarea,
.level-card--locked .stTextInput input,
.level-card--locked div[data-testid="stRadio"] {
    filter: grayscale(0.65);
    opacity: 0.8;
}

.level-card__lock-hint {
    display: flex;
    align-items: center;
    gap: 0.65rem;
    font-size: 0.86rem;
    color: rgba(42, 55, 78, 0.88);
    background: rgba(64, 84, 114, 0.12);
    border: 1px dashed rgba(64, 84, 114, 0.35);
    border-radius: 14px;
    padding: 0.7rem 0.85rem;
    margin-bottom: 1.05rem;
}

.level-card__lock-hint strong {
    color: rgba(32, 45, 68, 0.92);
}

.question-block__hint {
    margin-top: 0.4rem;
    background: rgba(255, 193, 99, 0.16);
    border-left: 4px solid rgba(255, 166, 43, 0.5);
    padding: 0.55rem 0.75rem;
    border-radius: 10px;
    font-size: 0.85rem;
    color: rgba(132, 77, 7, 0.92);
}

.question-block__warning {
    margin-top: 0.6rem;
    background: rgba(206, 104, 86, 0.14);
    border-left: 5px solid rgba(206, 104, 86, 0.9);
    padding: 0.7rem 0.85rem;
    border-radius: 10px;
    font-size: 0.86rem;
    color: rgba(122, 36, 24, 0.95);
}

.question-block__error {
    margin-top: 0.35rem;
    font-size: 0.78rem;
    color: rgba(171, 44, 38, 0.98);
    font-weight: 600;
}

.stepper-form__counter {
    text-align: right;
    font-size: 0.75rem;
    color: var(--text-500);
    margin-top: -0.4rem;
}

.stepper-form__counter--alert {
    color: #a35a00;
    font-weight: 600;
}

.stepper-form__hint {
    margin-top: 0.45rem;
    background: rgba(255, 193, 99, 0.16);
    border-left: 4px solid rgba(255, 166, 43, 0.5);
    padding: 0.55rem 0.75rem;
    border-radius: 10px;
    font-size: 0.86rem;
    color: rgba(132, 77, 7, 0.92);
}

.stepper-form__warning {
    margin-top: 0.6rem;
    background: rgba(206, 104, 86, 0.14);
    border-left: 5px solid rgba(206, 104, 86, 0.9);
    padding: 0.7rem 0.9rem;
    border-radius: 10px;
    font-size: 0.87rem;
    color: rgba(122, 36, 24, 0.95);
}

div[data-testid="stDataFrame"],
div[data-testid="stDataEditor"] {
    border: 1px solid rgba(var(--shadow-color), 0.12);
    border-radius: 6px;
    overflow: hidden;
    box-shadow: none;
    background: #ffffff;
}

div[data-testid="stDataFrame"] div[role="columnheader"],
div[data-testid="stDataEditor"] div[role="columnheader"] {
    background: #f5f7f9 !important;
    color: var(--text-700) !important;
    font-weight: 600;
    font-size: 0.88rem;
    text-transform: none;
    letter-spacing: 0;
    border-bottom: 1px solid rgba(12, 32, 20, 0.15);
    box-shadow: none;
}

div[data-testid="stDataFrame"] div[role="gridcell"],
div[data-testid="stDataEditor"] div[role="gridcell"] {
    color: var(--text-700);
    font-size: 0.9rem;
    border-bottom: 1px solid rgba(var(--forest-700), 0.1);
    border-right: 1px solid rgba(var(--forest-700), 0.08);
    padding: 0.45rem 0.6rem;
    background: #ffffff;
    word-wrap: break-word;
    white-space: normal;
    max-width: 300px;
}

div[data-testid="stDataFrame"] div[role="row"],
div[data-testid="stDataEditor"] div[role="row"] {
    transition: background 0.2s ease, box-shadow 0.2s ease;
}

div[data-testid="stDataFrame"] div[role="rowgroup"] > div:nth-child(odd) div[role="row"],
div[data-testid="stDataEditor"] div[role="rowgroup"] > div:nth-child(odd) div[role="row"] {
    background: #ffffff;
}

div[data-testid="stDataFrame"] div[role="rowgroup"] > div:nth-child(even) div[role="row"],
div[data-testid="stDataEditor"] div[role="rowgroup"] > div:nth-child(even) div[role="row"] {
    background: #fafbfc;
}

div[data-testid="stDataFrame"] div[role="rowgroup"] > div div[role="row"]:hover,
div[data-testid="stDataEditor"] div[role="rowgroup"] > div div[role="row"]:hover {
    background: #f3f5f7;
    box-shadow: none;
}

div[data-testid="stDataFrame"] div[role="rowgroup"] > div div[role="row"]:hover div[role="gridcell"],
div[data-testid="stDataEditor"] div[role="rowgroup"] > div div[role="row"]:hover div[role="gridcell"] {
    border-bottom-color: transparent;
}
</style>
""",
    unsafe_allow_html=True,
)

# Banner institucional de la Hoja de IRL (debe ir al principio)
render_irl_banner()

# ========================================
# SELECTOR DE MODO: CONECTADO vs INDIVIDUAL
# ========================================
st.markdown("---")
st.subheader("🔀 Modo de Trabajo")

# Selector de modo con componentes nativos
col_mode_info1, col_mode_info2 = st.columns(2)

with col_mode_info1:
    st.info("""
    **🔗 Modo Conectado**
    - Usa proyectos de Fase 0
    - Solo proyectos priorizados
    - Validación automática
    - Flujo continuo
    """)

with col_mode_info2:
    st.info("""
    **🔓 Modo Individual**
    - Todos los proyectos
    - Sin depender de ranking
    - Carga archivos directo
    - Máxima flexibilidad
    """)

# Selector de modo funcional
current_mode = st.session_state.get('irl_mode', 'conectado')
mode_irl = st.radio(
    "Selecciona tu modo:",
    options=["🔗 Modo Conectado", "🔓 Modo Individual"],
    index=0 if current_mode == 'conectado' else 1,
    horizontal=True,
    key='radio_irl_mode'
)

# Guardar modo seleccionado
if "🔗" in mode_irl:
    st.session_state.irl_mode = 'conectado'
else:
    st.session_state.irl_mode = 'individual'

# Status y validación
if st.session_state.irl_mode == 'conectado':
    payload = st.session_state.get('fase1_payload')
    fase1_ready = st.session_state.get('fase1_ready', False)
    
    if not (payload and fase1_ready):
        st.error("⚠️ **Modo Conectado requiere datos de Fase 0**")
        
        if st.button('📂 Ir a Fase 0 para calcular ranking', type="primary", use_container_width=True):
            fase0_page = next(Path("pages").glob("02_*_Fase_0_Portafolio.py"), None)
            if fase0_page:
                st.switch_page(str(fase0_page))
        st.stop()
    else:
        ranking_df = payload.get('ranking', pd.DataFrame())
        num_proyectos = len(ranking_df) if not ranking_df.empty else 0
        st.success(f"✅ Modo Conectado activo - {num_proyectos} proyecto(s) disponible(s)")
else:
    # Tips para modo individual
    with st.expander("💡 Tips para Modo Individual", expanded=True):
        st.markdown("""
        - ✅ Descarga la plantilla Excel con todas las preguntas
        - ✅ Completa solo las respuestas VERDADERAS (ahorra tiempo)
        - ✅ Sube el archivo cuando termines
        - ✅ Puedes exportar y consolidar después
        """)

st.markdown("---")

# ========================================
# LÓGICA SEGÚN MODO SELECCIONADO
# ========================================

fase0_page = next(Path("pages").glob("02_*_Fase_0_Portafolio.py"), None)
fase2_page = next(Path("pages").glob("04_*_Fase_2_*.py"), None)

if st.session_state.irl_mode == 'conectado':
    # MODO CONECTADO: Requiere payload de Fase 0
    payload = st.session_state.get('fase1_payload')
    fase1_ready = st.session_state.get('fase1_ready', False)
    
    if "fase2_ready" not in st.session_state:
        st.session_state["fase2_ready"] = False

    if not payload or not fase1_ready:
        st.warning('Calcula el ranking de candidatos en Fase 0 y usa el boton "Ir a Fase 1" para continuar.')
        if fase0_page:
            if st.button('Ir a Fase 0', key='btn_ir_fase0_desde_fase1'):
                st.switch_page(str(fase0_page))
        st.stop()

    ranking_df = payload['ranking'].copy().reset_index(drop=True)
    if ranking_df.empty:
        st.warning('El ranking recibido esta vacio. Recalcula la priorizacion en Fase 0.')
        if fase0_page:
            if st.button('Recalcular en Fase 0', key='btn_recalcular_fase0'):
                st.switch_page(str(fase0_page))
        st.stop()

    metrics_cards = payload.get('metrics_cards', [])
    umbrales = payload.get('umbrales', {})
        
    # Ocultamos visualización de métricas para un layout más compacto
    # (se conserva la data para otros cálculos).

    # Se elimina la sección visual de “Ranking de candidatos priorizados” para simplificar la interfaz.
    ranking_keys = ranking_df[['id_innovacion', 'ranking']].copy()
    ranking_keys['id_str'] = ranking_keys['id_innovacion'].astype(str)

    df_port = utils.normalize_df(db.fetch_df())
    df_port['id_str'] = df_port['id_innovacion'].astype(str)
    df_port = df_port[df_port['id_str'].isin(ranking_keys['id_str'])].copy()
    if df_port.empty:
        st.warning('Los proyectos del ranking ya no estan disponibles en el portafolio maestro. Recalcula la priorizacion en Fase 0.')
        if fase0_page:
            if st.button('Volver a Fase 0', key='btn_volver_recalcular'):
                st.switch_page(str(fase0_page))
        st.stop()

    order_map = dict(zip(ranking_keys['id_str'], ranking_keys['ranking']))
    df_port['orden_ranking'] = df_port['id_str'].map(order_map)
    df_port = df_port.sort_values('orden_ranking').reset_index(drop=True)
    df_port = df_port.drop(columns=['id_str', 'orden_ranking'], errors='ignore')

else:
    # MODO INDIVIDUAL: Trabaja sin payload de Fase 0
    if "fase2_ready" not in st.session_state:
        st.session_state["fase2_ready"] = False
    
    # Obtener todos los proyectos disponibles del portafolio maestro
    df_port = utils.normalize_df(db.fetch_df())
    
    if df_port.empty:
        st.warning('⚠️ No hay proyectos en el portafolio maestro. Carga proyectos en Fase 0 primero.')
        if fase0_page:
            if st.button('Ir a Fase 0', key='btn_ir_fase0_individual'):
                st.switch_page(str(fase0_page))
        st.stop()
    
    # En modo individual no hay ranking, trabajamos con todos los proyectos
    payload = None
    ranking_df = pd.DataFrame()
    metrics_cards = []
    umbrales = {}



def parse_project_id(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        st.error('No se puede registrar la evaluacion porque el identificador del proyecto no es numerico. Revisa la Fase 0.')
        st.stop()


def fmt_opt(identificador: int) -> str:
    fila = df_port.loc[df_port["id_innovacion"] == identificador]
    if fila.empty:
        return str(identificador)
    return f"{identificador} - {fila['nombre_innovacion'].values[0]}"


with st.container():
    st.markdown("<div class='section-shell'>", unsafe_allow_html=True)
    st.markdown("### Selecciona un proyecto del portafolio maestro")
    ids = df_port["id_innovacion"].tolist()
    seleccion = st.selectbox("Proyecto", ids, format_func=fmt_opt)
    st.markdown("</div>", unsafe_allow_html=True)


project_id = parse_project_id(seleccion)

previous_project_id = st.session_state.get("fase2_last_project_id")
if previous_project_id is not None and previous_project_id != project_id:
    st.session_state["fase2_ready"] = False
    st.session_state.pop("fase2_payload", None)
st.session_state["fase2_last_project_id"] = project_id

selected_project = df_port.loc[df_port["id_innovacion"] == project_id].iloc[0]
impacto_txt = selected_project.get("impacto") or "No informado"
estado_txt = selected_project.get("estatus") or "Sin estado"
responsable_txt = selected_project.get("responsable_innovacion") or "Sin responsable asignado"
transferencia_txt = selected_project.get("potencial_transferencia") or "Sin potencial declarado"
evaluacion_val = selected_project.get("evaluacion_numerica")
evaluacion_txt = f"{float(evaluacion_val):.1f}" if pd.notna(evaluacion_val) else "—"

project_snapshot = {
    "id_innovacion": int(project_id),
    "nombre_innovacion": selected_project.get("nombre_innovacion", ""),
    "potencial_transferencia": transferencia_txt,
    "impacto": impacto_txt,
    "estatus": estado_txt,
    "responsable_innovacion": responsable_txt,
    "evaluacion_numerica": float(evaluacion_val) if pd.notna(evaluacion_val) else None,
}

selection_meta = [
    ("Impacto estratégico", impacto_txt),
    ("Estado actual", estado_txt),
    ("Responsable de innovación", responsable_txt),
    ("Evaluación Fase 0", evaluacion_txt),
]

meta_items_html = "".join(
    f"<div class='selection-card__meta-item'>"
    f"<span class='selection-card__meta-label'>{escape(label)}</span>"
    f"<span class='selection-card__meta-value'>{escape(str(value))}</span>"
    "</div>"
    for label, value in selection_meta
)

selection_card_html = f"""
<div class='selection-card'>
    <span class='selection-card__badge'>Proyecto seleccionado</span>
    <h3 class='selection-card__title'>{escape(selected_project['nombre_innovacion'])}</h3>
    <p class='selection-card__subtitle'>{escape(str(transferencia_txt))}</p>
    <div class='selection-card__meta'>
        {meta_items_html}
    </div>
</div>
"""

with st.container():
    st.markdown("<div class='section-shell'>", unsafe_allow_html=True)
    st.markdown(selection_card_html, unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

with st.container():
    st.markdown("<div class='section-shell'>", unsafe_allow_html=True)
    st.markdown("### 📊 Evaluación IRL - Flujo de Trabajo")
    
    # Guía visual simple con columnas
    col_paso1, col_paso2, col_paso3, col_paso4 = st.columns(4)
    
    with col_paso1:
        st.markdown("### 1️⃣")
        st.markdown("**📥 Descargar**")
        st.caption("Plantilla Excel pre-llenada")
    
    with col_paso2:
        st.markdown("### 2️⃣")
        st.markdown("**📝 Completar**")
        st.caption("Cambia a VERDADERO")
    
    with col_paso3:
        st.markdown("### 3️⃣")
        st.markdown("**📤 Subir**")
        st.caption("Carga el archivo")
    
    with col_paso4:
        st.markdown("### 4️⃣")
        st.markdown("**✅ Confirmar**")
        st.caption("Aplica al sistema")
    
    st.markdown("---")
    
    # Sección de descarga
    st.markdown("#### 📥 Paso 1: Descargar Plantilla de Evaluación")
    
    col_download, col_info = st.columns([1, 2])
    
    with col_download:
        excel_template = generate_irl_excel_template()
        st.download_button(
            label="⬇️ Descargar Plantilla Excel",
            data=excel_template,
            file_name=f"Evaluacion_IRL_Proyecto_{project_id}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
    
    with col_info:
        st.info("""
        **📋 La plantilla incluye:**
        - ✅ Todas las respuestas PRE-LLENADAS en FALSO
        - ✅ Solo cambias a VERDADERO las que SÍ cumplan
        - ✅ Agrega evidencias en las VERDADERO
        - ✅ Ahorra 80% del tiempo de evaluación
        """)
    
    st.markdown("---")
    
    # Paso 2: Subir archivo completado
    st.markdown("#### 📤 Paso 2: Subir Evaluación Completada")
    
    uploaded_file = st.file_uploader(
        "Selecciona el archivo Excel completado",
        type=["xlsx"],
        help="Sube la plantilla que descargaste y completaste offline"
    )
    
    # Almacenar respuestas pendientes y generar tabla de revisión
    if uploaded_file is not None:
        if 'irl_excel_file_loaded' not in st.session_state or st.session_state.get('irl_excel_file_loaded') != uploaded_file.name:
            responses = load_irl_excel_responses(uploaded_file)
            if responses:
                st.session_state.pending_irl_responses = responses
                st.session_state.irl_excel_file_loaded = uploaded_file.name
                
                # Generar datos para la tabla de revisión
                revision_data = []
                for key, value in responses.items():
                    if key.startswith('resp_'):
                        # Extraer información del key: resp_DIMENSION_NIVEL_PREGUNTA
                        parts = key.replace('resp_', '').split('_')
                        dimension = parts[0]
                        nivel = parts[1]
                        pregunta_num = parts[2]
                        
                        # Obtener evidencia correspondiente
                        evid_key = f"evid_{dimension}_{nivel}_{pregunta_num}"
                        evidencia = responses.get(evid_key, "")
                        
                        # Buscar texto de la pregunta
                        levels = LEVEL_DEFINITIONS.get(dimension, [])
                        pregunta_texto = ""
                        for level in levels:
                            if level["nivel"] == int(nivel):
                                preguntas = level.get("preguntas", [])
                                if int(pregunta_num) <= len(preguntas):
                                    pregunta_texto = preguntas[int(pregunta_num) - 1]
                                break
                        
                        dim_desc = DIMENSION_DESCRIPTIONS.get(dimension, dimension)
                        
                        revision_data.append({
                            'Dimensión': f"{dimension} - {dim_desc}",
                            'Nivel': nivel,
                            'Pregunta #': pregunta_num,
                            'Pregunta': pregunta_texto[:80] + "..." if len(pregunta_texto) > 80 else pregunta_texto,
                            'Respuesta': value,
                            'Evidencia': evidencia[:60] + "..." if len(evidencia) > 60 else evidencia if evidencia else "(Sin evidencia)"
                        })
                
                st.session_state.irl_revision_data = revision_data
    
    # Botones de acción para archivo revisado
    if st.session_state.get('irl_excel_file_loaded') and 'pending_irl_responses' in st.session_state:
        st.markdown("---")
        st.markdown("##### ✅ Confirmar y Aplicar")
        
        col_aplicar, col_cancelar = st.columns([1, 1])
        
        with col_aplicar:
            if st.button("✅ Aplicar respuestas al sistema", use_container_width=True, type="primary"):
                # Aplicar todas las respuestas al session_state
                for key, value in st.session_state.pending_irl_responses.items():
                    st.session_state[key] = value
                
                # Inicializar estado y actualizar estados de niveles basados en respuestas
                _init_irl_state()
                _update_level_states_from_responses()
                
                # Calcular scores después de actualizar estados
                _sync_all_scores()
                
                # Marcar como aplicado
                st.session_state.irl_responses_applied = True
                
                # Limpiar pendientes pero mantener datos de revisión
                del st.session_state.pending_irl_responses
                
                st.success("✅ Respuestas aplicadas y niveles calculados correctamente.")
                st.rerun()
        
        with col_cancelar:
            if st.button("❌ Cancelar y subir otro archivo", use_container_width=True):
                # Limpiar todo sin aplicar
                if 'pending_irl_responses' in st.session_state:
                    del st.session_state.pending_irl_responses
                if 'irl_excel_file_loaded' in st.session_state:
                    del st.session_state.irl_excel_file_loaded
                if 'irl_revision_data' in st.session_state:
                    del st.session_state.irl_revision_data
                
                st.info("Archivo cancelado. Puedes subir un nuevo archivo.")
                st.rerun()
    
    st.markdown("---")
    
    # Paso 3: Mostrar resultados automáticamente después de aplicar
    if st.session_state.get('irl_responses_applied', False):
        st.markdown("### 📊 Resultados del Análisis IRL por Dimensión")
        st.caption("*Cada dimensión se evalúa de forma independiente (Niveles 0-9)*")
        
        # Asegurar que los estados están actualizados antes de calcular
        _init_irl_state()
        _update_level_states_from_responses()
        _sync_all_scores()
        
        st.markdown("---")
        
        # Calcular métricas por dimensión
        dimension_results = []
        for dimension, dim_desc in IRL_DIMENSIONS:
            nivel_alcanzado = st.session_state["irl_scores"].get(dimension, 0)
            counts = _compute_dimension_counts(dimension)
            porcentaje = (nivel_alcanzado / 9) * 100 if nivel_alcanzado > 0 else 0
            
            dimension_results.append({
                'dimension': dimension,
                'descripcion': dim_desc,
                'nivel': nivel_alcanzado,
                'porcentaje': porcentaje,
                'completado': counts['completed'],
                'total': counts['total']
            })
        
        # Tarjetas de dimensiones y gráfico radar en el mismo nivel
        st.markdown("#### 🎯 Resultados por Dimensión")
        
        # Crear layout: 6 tarjetas compactas + radar
        col_tarjetas, col_radar = st.columns([2, 1])
        
        with col_tarjetas:
            # Crear 6 columnas (una por dimensión) - 2 filas de 3
            cols_row1 = st.columns(3)
            cols_row2 = st.columns(3)
            
            for idx, result in enumerate(dimension_results):
                # Usar primera o segunda fila según el índice
                cols = cols_row1 if idx < 3 else cols_row2
                col_idx = idx % 3
                
                with cols[col_idx]:
                    # Colores profesionales más sutiles
                    if result['porcentaje'] >= 70:
                        color_principal = "#1565c0"  # Azul profesional oscuro
                        bgcolor = "#e3f2fd"  # Azul muy claro
                        icono = "✓"
                        icono_color = "#1976d2"
                    elif result['porcentaje'] >= 40:
                        color_principal = "#f57c00"  # Naranja
                        bgcolor = "#fff3e0"
                        icono = "◐"
                        icono_color = "#fb8c00"
                    else:
                        color_principal = "#757575"  # Gris profesional
                        bgcolor = "#f5f5f5"
                        icono = "○"
                        icono_color = "#9e9e9e"
                    
                    # Tarjeta compacta más profesional
                    st.markdown(f"""
                        <div style="background: linear-gradient(135deg, {bgcolor} 0%, white 100%); 
                                    border-left: 4px solid {color_principal}; 
                                    border-radius: 8px; padding: 0.7rem; margin-bottom: 0.5rem;
                                    box-shadow: 0 1px 4px rgba(0,0,0,0.06);">
                            <div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 0.4rem;">
                                <div style="font-weight: 600; color: {color_principal}; font-size: 0.9rem;">
                                    {result['dimension']}
                                </div>
                                <div style="font-size: 1.2rem; color: {icono_color}; font-weight: bold;">
                                    {icono}
                                </div>
                            </div>
                            <div style="text-align: center; background: white; border-radius: 6px; padding: 0.5rem; margin: 0.3rem 0;">
                                <div style="font-size: 1.6rem; font-weight: bold; color: {color_principal};">
                                    {result['nivel']}
                                </div>
                                <div style="font-size: 0.65rem; color: #999; text-transform: uppercase; letter-spacing: 0.5px;">
                                    de 9 niveles
                                </div>
                            </div>
                            <div style="text-align: center; font-size: 0.75rem; color: #666; margin-top: 0.3rem;">
                                {result['porcentaje']:.0f}% progreso
                            </div>
                        </div>
                    """, unsafe_allow_html=True)
        
        with col_radar:
            st.markdown("##### 📊 Visualización")
            
            # Crear gráfico radar
            labels = [r['dimension'] for r in dimension_results]
            values = [r['nivel'] for r in dimension_results]
            values_cycle = values + values[:1]
            theta = labels + labels[:1]
            
            radar_fig = go.Figure()
            radar_fig.add_trace(
                go.Scatterpolar(
                    r=values_cycle,
                    theta=theta,
                    fill="toself",
                    name="Nivel IRL",
                    line=dict(color="#1565c0", width=2.5),
                    fillcolor="rgba(21, 101, 192, 0.25)",
                    marker=dict(size=7, color="#1976d2"),
                    hovertemplate="<b>%{theta}</b><br>Nivel: %{r}/9<extra></extra>"
                )
            )
            
            radar_fig.update_layout(
                polar=dict(
                    radialaxis=dict(
                        visible=True,
                        range=[0, 9],
                        tickmode='linear',
                        tick0=0,
                        dtick=1,
                        gridcolor="rgba(0,0,0,0.08)",
                        gridwidth=1
                    ),
                    angularaxis=dict(
                        gridcolor="rgba(0,0,0,0.08)",
                        gridwidth=1
                    ),
                    bgcolor="rgba(255,255,255,0.9)"
                ),
                template="plotly_white",
                margin=dict(l=60, r=60, t=20, b=20),
                height=400,
                showlegend=False,
                font=dict(size=10, family="Arial, sans-serif")
            )
            
            st.plotly_chart(radar_fig, use_container_width=True)
        
        st.markdown("---")
        
        # Expander con detalle completo de preguntas/respuestas
        with st.expander("📋 Ver Detalle Completo de Respuestas por Dimensión", expanded=False):
            tab_labels = [f"{r['dimension']} - {r['descripcion']}" for r in dimension_results]
            tabs = st.tabs(tab_labels)
            
            for idx, result in enumerate(dimension_results):
                with tabs[idx]:
                    dimension = result['dimension']
                    levels = LEVEL_DEFINITIONS.get(dimension, [])
                    
                    for level in levels:
                        level_id = level["nivel"]
                        descripcion = level.get("descripcion", "")
                        preguntas = level.get("preguntas", [])
                        
                        st.markdown(f"**Nivel {level_id}**: {descripcion}")
                        
                        # Mostrar respuestas
                        for idx_p, pregunta in enumerate(preguntas, start=1):
                            resp_key = f"resp_{dimension}_{level_id}_{idx_p}"
                            evid_key = f"evid_{dimension}_{level_id}_{idx_p}"
                            
                            respuesta = st.session_state.get(resp_key, "FALSO")
                            evidencia = st.session_state.get(evid_key, "")
                            
                            icon = "✅" if respuesta == "VERDADERO" else "❌"
                            color = "#2e7d32" if respuesta == "VERDADERO" else "#d32f2f"
                            
                            evidencia_html = f"<br><em>Evidencia:</em> {evidencia}" if evidencia else ""
                            st.markdown(f"""
                                <div style="background: rgba(0,0,0,0.02); padding: 0.8rem; 
                                            border-left: 4px solid {color}; margin: 0.5rem 0; border-radius: 4px;">
                                    <strong>{icon} Pregunta {idx_p}:</strong> {pregunta}<br>
                                    <em>Respuesta:</em> <strong style="color: {color};">{respuesta}</strong>
                                    {evidencia_html}
                                </div>
                            """, unsafe_allow_html=True)
                        
                        st.markdown("---")
        
        # Botón para limpiar y volver a evaluar
        st.markdown("---")
        col_clear, col_empty = st.columns([1, 2])
        with col_clear:
            if st.button("🔄 Nueva Evaluación", use_container_width=True):
                # Limpiar todas las respuestas
                keys_to_delete = [k for k in st.session_state.keys() if k.startswith(('resp_', 'toggle_', 'evid_'))]
                for k in keys_to_delete:
                    del st.session_state[k]
                
                # Limpiar flags
                st.session_state.irl_responses_applied = False
                if 'irl_excel_file_loaded' in st.session_state:
                    del st.session_state.irl_excel_file_loaded
                
                st.rerun()
    
    st.markdown("</div>", unsafe_allow_html=True)

with st.container():
    st.markdown("<div class='section-shell'>", unsafe_allow_html=True)
    df_respuestas = _collect_dimension_responses()
    detalles_dimensiones = _collect_dimension_details()
    with st.expander('Detalle de niveles por dimension', expanded=False):
        st.markdown("<div class='details-panel'>", unsafe_allow_html=True)
        if df_respuestas.empty:
            st.info("Aún no hay niveles respondidos en esta evaluación.")

        if detalles_dimensiones:
            tab_labels = [
                f"{info['label']}" if info["label"] else dimension
                for dimension, info in detalles_dimensiones.items()
            ]
            st.markdown("**Preguntas y respuestas por dimensión**")
            tabs = st.tabs(tab_labels)
            for idx, (dimension, info) in enumerate(detalles_dimensiones.items()):
                with tabs[idx]:
                    detalle_df = pd.DataFrame(info["rows"])
                    if detalle_df.empty:
                        st.info("No hay niveles configurados para esta dimensión.")
                    else:
                        render_table(
                            detalle_df,
                            key=f'fase1_detalle_dimensiones_{dimension}',
                            include_actions=False,
                            hide_index=True,
                            page_size_options=(10, 25, 50),
                            default_page_size=10,
                        )
        else:
            st.warning("No se encontraron definiciones de niveles para las dimensiones IRL.")
    st.markdown("</div>", unsafe_allow_html=True)
    # Use cached puntaje when available; avoid recalculating on each rerun to improve responsiveness.
    puntaje = st.session_state.get("irl_last_puntaje")
    if puntaje is None and df_respuestas.empty:
        puntaje = None

    col_guardar, col_ayuda = st.columns([1, 1])
    with col_guardar:
        finalize_clicked = st.button("Finalizar evaluación", type="primary")
        if finalize_clicked:
            # If we don't have a cached puntaje, compute it now (user-triggered expensive op)
            if puntaje is None and not df_respuestas.empty:
                try:
                    computed = trl.calcular_trl(df_respuestas[["dimension", "nivel", "evidencia"]])
                    st.session_state["irl_last_puntaje"] = computed
                    puntaje = computed
                except Exception:
                    st.session_state["irl_last_puntaje"] = None
                    puntaje = None

            if puntaje is None:
                st.info(
                    "La evaluación se guardará sin campos respondidos ni niveles acreditados para que puedas avanzar a la fase 2."
                )
            trl_value = float(puntaje) if puntaje is not None else None
            try:
                save_trl_result(
                    project_id,
                    df_respuestas[["dimension", "nivel", "evidencia"]],
                    trl_value,
                )
                _sync_all_scores()
                historial = get_trl_history(project_id)
                fecha_eval = historial["fecha_eval"].iloc[0] if not historial.empty else None
                responses_records = df_respuestas[["dimension", "nivel", "evidencia"]].to_dict("records")
                st.session_state["fase2_ready"] = True
                st.session_state["fase2_payload"] = {
                    "project_id": project_snapshot["id_innovacion"],
                    "project_snapshot": project_snapshot.copy(),
                    "responses": responses_records,
                    "irl_score": trl_value,
                    "fecha_eval": fecha_eval,
                }
                st.session_state["fase2_last_project_id"] = project_id
                if fase2_page:
                    st.switch_page(str(fase2_page))
                else:
                    st.success("Evaluacion guardada correctamente.")
                    _rerun_app()
            except Exception as error:
                st.error(f"Error al guardar: {error}")

        fase2_payload = st.session_state.get("fase2_payload")
        fase2_ready_for_project = (
            st.session_state.get("fase2_ready", False)
            and fase2_payload
            and fase2_payload.get("project_id") == project_id
        )
        go_fase2 = st.button(
            "Ir a Fase 2",
            key="btn_go_fase2",
            disabled=not fase2_ready_for_project,
        )
        if go_fase2:
            if fase2_page:
                st.switch_page(str(fase2_page))
            else:
                st.warning("La página de Fase 2 aún no está disponible.")

    with col_ayuda:
        st.info(
            "El guardado crea un registro por dimensión con las evidencias acreditadas y asocia el IRL global a la misma fecha de evaluación."
        )
    st.markdown("</div>", unsafe_allow_html=True)
